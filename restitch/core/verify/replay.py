"""The independent verifier — every rule re-derived, nothing trusted.

Ported from the source project's verifier and generalized. Positions come from
the caller's independently loaded inputs; movement lines come from the saved
workbook's own cells; every threshold comes from the Policy; the geography
predicate and the localization optimality re-solve are both geo-constrained —
the optimality proof solves the same scoped problem allocation did.

Disabled rules surface as SKIPPED verifications in place, mirroring the build
battery: the count never silently shrinks.
"""
from __future__ import annotations

from collections import Counter, defaultdict

import openpyxl

from ..checks import Battery
from ..ids import store_id
from ..model import canon_city, make_geo_edge, make_season_age, pair_km
from ..policy import Policy
from ..rules import RULES

# verify id -> owning rule id, DERIVED from the registry (absent = kernel).
_V_RULE = {vid: r.id for r in RULES for vid in r.verify_ids}

SEC = "independent verify"


def _wave2_sheet(pol) -> str:
    return f"1c · Wave 2 — {'·'.join(pol.wave2_pivotals)}"[:31]


# Columns the verifier reads, BY NAME — positional parsing once let a
# rearranged layout mis-verify silently. Money headers carry an optional
# currency suffix, so matching is prefix-based for those two.
_NEED_HDRS = ("Priority", "Program", "Take from", "Deliver to", "Style", "Cat",
              "Size", "Qty", "RoS from", "RoS to", "Confidence")


def _hmap(hdr_row, sheet) -> dict:
    idx = {str(h).strip(): j for j, h in enumerate(hdr_row) if h is not None}
    missing = [h for h in _NEED_HDRS if h not in idx]
    assert not missing, f"sheet {sheet!r}: required headers missing: {missing}"
    return idx


def parse_lines(xlsx, pol) -> list:
    """Movement lines from the saved sheets' own cells, columns resolved by
    header NAME — a rearranged layout refuses loudly instead of mis-parsing."""
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    lines = []
    for sheet, tag in (("1 · Movement — Take & Deliver", "inter"),
                       ("1b · Same-City Moves", "intra"),
                       (_wave2_sheet(pol), "wave2"),
                       ("1d · Assembly — New Sets", "asm")):
        ws = wb[sheet]
        ix = None
        for row in ws.iter_rows(values_only=True):
            if ix is None:
                if row and row[0] == "Priority":
                    ix = _hmap(row, sheet)
                continue
            if row[ix["Priority"]] not in ("P1", "P2", "P3"):
                continue

            def g(name, r=row, ix=ix):
                j = ix[name]
                return r[j] if j < len(r) else None

            lines.append(dict(band=g("Priority"), program=g("Program"),
                              src=store_id(g("Take from")),
                              dst=store_id(g("Deliver to")),
                              style=str(g("Style")).strip(), cat=g("Cat"),
                              size=str(g("Size")), qty=float(g("Qty") or 0),
                              sros=float(g("RoS from") or 0),
                              dros=float(g("RoS to") or 0),
                              conf=g("Confidence"), sheet=tag))
        assert ix is not None, f"sheet {sheet!r} not parsed"
    assert all(x["program"] == "HEAL3" for x in lines if x["sheet"] == "wave2") and \
           all(x["program"] != "HEAL3" for x in lines if x["sheet"] != "wave2"), \
        "wave-2 lines must live on their sheet and only there"
    assert all(x["program"] == "ASM" for x in lines if x["sheet"] == "asm") and \
           all(x["program"] != "ASM" for x in lines if x["sheet"] != "asm"), \
        "assembly lines must live on sheet 1d and only there"
    return lines


def verify_workbook(xlsx, inputs, pol: Policy, labels=None) -> Battery:
    """Replay the saved workbook against independently loaded inputs.
    `labels` mirrors the renderer's dict — the recon parse anchors on the
    tenant's own store-id header."""
    sid_hdr = (labels or {}).get("store_id", "Store ID")
    stores, holdings = inputs.stores, inputs.holdings
    norms, naw = inputs.norms, inputs.norms_context
    b = Battery()
    rules_on = {r.id: bool(r.enabled(pol)) for r in RULES}

    def check(vid, name, cond, detail):
        b.add(vid, SEC, name, bool(cond), detail)
        assert cond, f"VERIFY FAIL — {vid} {name}: {detail}"

    def gate(vid, name):
        owner = _V_RULE.get(vid)
        if owner is not None and not rules_on[owner]:
            b.skip(vid, SEC, name, "rule disabled by policy")
            return False
        return True

    # ── positions re-derived from the raw inputs alone ──
    categories = sorted({c for (_sh, c) in norms} | {h.cat for h in holdings.values()})
    soh, soh_cat, soh_pss, held = Counter(), Counter(), Counter(), Counter()
    for (sh, st), h in holdings.items():
        for z, q in h.sizes.items():
            if q > 0:
                soh[sh] += q
                soh_cat[(sh, h.cat)] += q
                soh_pss[(sh, st, str(z))] += q
                held[(sh, st)] += q
    plants = set(soh)

    def norm_of(sh):
        return sum(norms.get((sh, c), 0.0) for c in categories)

    aliases = dict(pol.city_aliases)
    m_city = {sh: canon_city(s.city, aliases) for sh, s in stores.items()}
    nn = {sh for sh in plants if norm_of(sh) <= 0} - set(pol.afs_stores)
    frozen_m = {sh for sh, s in stores.items()
                if s.blocked or (s.status1 and s.status1 != "Running")} \
        | set(pol.foco_frozen)
    if not pol.evac_no_norm:
        # containment mirror (review M4): with evacuation off, normless doors
        # freeze in place — the verifier re-derives the same set from norms
        frozen_m |= nn
    m_new = set()
    if pol.new_door_donor_freeze and pol.new_door_cutoff is not None:
        for sh, s in stores.items():
            if (s.opened is not None and s.opened >= pol.new_door_cutoff
                    or s.opened is None and s.l2l in pol.new_door_l2l):
                m_new.add(sh)
        m_new -= frozen_m | set(pol.afs_stores)

    lines = parse_lines(xlsx, pol)
    tot = sum(x["qty"] for x in lines)
    W2 = tuple(pol.wave2_pivotals)

    # 1 · executable store→store inside the live network
    bad = [x for x in lines if x["src"] not in plants or x["dst"] not in plants
           or x["src"] == x["dst"] or x["qty"] < 1 or x["qty"] != int(x["qty"])]
    check("v_exec", "Every line is executable store→store inside the live network",
          not bad, f"{len(lines):,} lines · {tot:,.0f} u · {len(bad)} bad")

    # 2 · frozen doors (both ways)
    if gate("v_afs", "Named frozen doors on zero lines"):
        afs = set(pol.afs_stores)
        afs_l = [x for x in lines if x["src"] in afs or x["dst"] in afs]
        check("v_afs", "Named frozen doors appear on ZERO lines", not afs_l,
              f"{len(afs_l)} lines touch one · "
              f"{sum(soh[s] for s in afs):,.0f} u stays exactly where it is")

    # 3 · frozen styles
    if gate("v_core", "Frozen stylecodes on zero lines"):
        core_l = [x for x in lines if pol.is_core(x["style"])]
        check("v_core", "Frozen stylecodes (base or variant) appear on ZERO lines",
              not core_l, f"{len(core_l)} frozen-style lines")

    # 4 · no-norm doors never receive
    if gate("v_no_norm_in", "No-norm doors receive nothing"):
        nn_in = [x for x in lines if x["dst"] in nn]
        check("v_no_norm_in",
              "No-norm doors (re-derived from the raw norms) receive NOTHING",
              not nn_in,
              f"{len(nn)} no-norm doors hold stock · {len(nn_in)} inbound lines · "
              f"{sum(x['qty'] for x in lines if x['src'] in nn):,.0f} u evacuated out")

    # 5 · conservation vs raw positions
    out_u, in_u = Counter(), Counter()
    for x in lines:
        out_u[x["src"]] += x["qty"]
        in_u[x["dst"]] += x["qty"]
    over = [sh for sh, o in out_u.items() if o > soh[sh] + 1e-6]
    check("v_conservation",
          "Conservation — out == in, and no door ships more than its raw SOH",
          abs(sum(out_u.values()) - sum(in_u.values())) < 1e-6 and not over,
          f"out {sum(out_u.values()):,.0f} == in {sum(in_u.values()):,.0f} · "
          f"{len(over)} over-shipped doors")

    # ── kernel replay — per-size ledger truths the ARTIFACT must obey ──
    # (added after review C1: the saved file once passed with a forged
    # fragment out of a Complete set and a ghost size shipped from zero)
    shipped = Counter()
    pos = Counter(soh_pss)
    posw1 = Counter(soh_pss)
    for x in lines:
        k_s = (x["src"], x["style"], x["size"])
        k_d = (x["dst"], x["style"], x["size"])
        shipped[k_s] += x["qty"]
        pos[k_s] -= x["qty"]
        pos[k_d] += x["qty"]
        if x["sheet"] in ("inter", "intra"):
            posw1[k_s] -= x["qty"]
            posw1[k_d] += x["qty"]

    oversz = [k for k, q in shipped.items() if q > soh_pss.get(k, 0) + 1e-6]
    check("v_overdraw_size",
          "PER-SIZE OVERDRAW — no door ships more of one style+size than its "
          "raw opening",
          not oversz, f"{len(shipped):,} ship keys · {len(oversz)} over-drawn"
          + (f" e.g. {oversz[0]}" if oversz else ""))

    neg = [k for k, q in pos.items() if q < -1e-9]
    check("v_no_negative",
          "NO NEGATIVE POSITIONS — every line applied leaves all replayed "
          "positions >= 0",
          not neg, f"{len(pos):,} positions · {len(neg)} negative")

    ships_k = {(x["src"], x["style"], x["size"]) for x in lines}
    recvs_k = {(x["dst"], x["style"], x["size"]) for x in lines}
    rt = ships_k & recvs_k
    check("v_round_trips",
          "NO ROUND TRIPS — no store both ships and receives the same "
          "style+size",
          not rt, f"{len(ships_k):,} ship × {len(recvs_k):,} receive keys · "
          f"{len(rt)} overlap")

    PIV = list(pol.pivotals)
    complete = {k for k, h in holdings.items()
                if all(h.sizes.get(p, 0) > 0 for p in PIV)}
    frag = [x for x in lines if x["program"] != "SET"
            and (x["src"], x["style"]) in complete]
    set_moved, set_dsts2 = Counter(), defaultdict(set)
    for x in lines:
        if x["program"] == "SET":
            set_moved[(x["src"], x["style"])] += x["qty"]
            set_dsts2[(x["src"], x["style"])].add(x["dst"])
    set_bad = [k for k, q in set_moved.items()
               if abs(q - held[k]) > 1e-9 or len(set_dsts2[k]) != 1]
    check("v_full_set_freeze",
          "FULL-SET FREEZE — a Complete holding (re-derived from raw SOH) "
          "sources only whole-set relocations, entire and to ONE door",
          not frag and not set_bad,
          f"{len(complete):,} Complete holdings · {len(frag)} fragment lines "
          f"from one · {len(set_moved):,} SET sources · {len(set_bad)} not "
          "whole-to-one-door")

    heal_w1 = {(x["dst"], x["style"]) for x in lines
               if x["sheet"] in ("inter", "intra") and x["program"] == "HEAL"}
    inc_heal = [k for k in heal_w1
                if any(pos[(k[0], k[1], p)] <= 1e-9 for p in PIV)]
    check("v_heal_completion",
          "HEAL COMPLETION — every wave-1 healed door ends holding every "
          "pivotal size",
          not inc_heal, f"{len(heal_w1):,} healed sets · {len(inc_heal)} incomplete")

    p1_u = sum(x["qty"] for x in lines if x["band"] == "P1")
    grp = defaultdict(set)
    for x in lines:
        if x["program"] in ("HEAL", "SET", "HEAL3", "ASM"):
            grp[(x["program"], x["dst"], x["style"])].add(x["band"])
    split = [k for k, v in grp.items() if len(v) > 1]
    check("v_budget",
          f"BUDGET — P1 within {pol.movement_budget:,} units; heal/relocation "
          "atoms never split across bands",
          p1_u <= pol.movement_budget + 1e-9 and not split,
          f"P1 {p1_u:,.0f} u of {pol.movement_budget:,} · {len(grp):,} "
          f"set-atoms · {len(split)} split")

    if pol.protect_selling_fragments:
        keep_bad = []
        for k, rec in inputs.perf.items():
            if rec.get("sales8", 0) <= 0:
                continue
            h = holdings.get(k)
            if h is None or k[0] in nn:
                continue          # evacuation waives the keep-front
            if sum(1 for p in PIV if h.sizes.get(p, 0) <= 0) < 3:
                continue          # the law binds on selling BROKEN holdings
            for p in PIV:
                if h.sizes.get(p, 0) >= 1 and pos[(k[0], k[1], p)] < 1 - 1e-9:
                    keep_bad.append((k, p))
        check("v_keep_fronts",
              "KEEP-FRONTS — selling Broken holdings keep 1 unit of each held "
              "pivotal",
              not keep_bad, f"{len(keep_bad)} stripped fronts")
    else:
        b.skip("v_keep_fronts", SEC, "Keep-fronts",
               "protect_selling_fragments disabled by policy")

    outc = Counter()
    for x in lines:
        outc[(x["src"], x["cat"])] += x["qty"]

    # 6 · transit floor
    if gate("v_transit", "Transit floor"):
        t_bad = []
        for sh, o in out_u.items():
            n = norm_of(sh)
            if n > 0 and soh[sh] - o < min(soh[sh], pol.transit_floor * n) - 1e-6:
                t_bad.append(("ALL", sh))
        for (sh, cat), o in outc.items():
            nc = norms.get((sh, cat), 0.0)
            if nc > 0 and soh_cat[(sh, cat)] - o < min(soh_cat[(sh, cat)],
                                                       pol.transit_floor * nc) - 1e-6:
                t_bad.append((cat, sh))
        check("v_transit",
              f"TRANSIT FLOOR {pol.transit_floor:.0%} — opening − all out never dips "
              "below, combined AND per category",
              not t_bad, f"{len(out_u)} donor doors · {len(t_bad)} below")

    # 7 · fill floor (end state)
    if gate("v_fill", "Fill floor"):
        inc = Counter()
        for x in lines:
            inc[(x["dst"], x["cat"])] += x["qty"]
        f_bad = []
        for sh in plants:
            n = norm_of(sh)
            if n > 0 and soh[sh] - out_u[sh] + in_u[sh] < \
                    min(soh[sh], pol.fill_floor * n) - 1e-6:
                f_bad.append(("ALL", sh))
        for (sh, cat), nc in norms.items():
            if nc <= 0:
                continue
            e = soh_cat[(sh, cat)] - outc[(sh, cat)] + inc[(sh, cat)]
            if e < min(soh_cat[(sh, cat)], pol.fill_floor * nc) - 1e-6:
                f_bad.append((cat, sh))
        check("v_fill",
              f"FILL FLOOR {pol.fill_floor:.0%} — END fill never below, combined AND "
              "per category", not f_bad, f"{len(f_bad)} doors pushed below")

    # 8 · receiver peak cap
    if gate("v_peak", "Receiver peak cap"):
        in_ch, in_w1, in_ch_w1 = Counter(), Counter(), Counter()
        for x in lines:
            if x["program"] not in ("HEAL", "HEAL3"):
                in_ch[x["dst"]] += x["qty"]
            if x["sheet"] in ("inter", "intra"):
                in_w1[x["dst"]] += x["qty"]
                if x["program"] not in ("HEAL", "HEAL3"):
                    in_ch_w1[x["dst"]] += x["qty"]
        p_bad = []
        heal_only = 0
        for sh, _iq in in_u.items():
            n = norm_of(sh)
            if n <= 0:
                continue
            if in_ch[sh] <= 0:
                heal_only += 1
            if in_ch_w1[sh] > 0 and soh[sh] + in_w1[sh] > \
                    max(soh[sh], pol.recv_peak_cap * n) + 1e-6:
                p_bad.append(("w1", sh))
            if in_ch[sh] > 0 and soh[sh] + in_ch[sh] > \
                    max(soh[sh], pol.recv_peak_cap * n) + 1e-6:
                p_bad.append(("chosen", sh))
        check("v_peak",
              f"RECEIVER PEAK CAP {pol.recv_peak_cap:.0%} — wave-1 strong form + "
              "chosen-only across the full plan; heals exempt",
              not p_bad, f"{len(in_u)} receiving doors · {len(p_bad)} breaches · "
              f"{heal_only} heal-only doors (exempt)")

    # 9 · RoS confidence from the sheet's own columns
    if gate("v_ros", "RoS confidence"):
        c_bad = []
        for x in lines:
            if x["program"] in ("HEAL", "HEAL3"):
                if x["dros"] < x["sros"] - 1e-9:
                    c_bad.append(x)
            elif x["program"] == "ASM":
                continue
            elif x["dros"] < max(pol.min_recv_ros,
                                 pol.ros_uplift * x["sros"]) - 1e-9:
                c_bad.append(x)
        cu = Counter()
        for x in lines:
            cu[x["conf"]] += x["qty"]
        check("v_ros",
              "RoS CONFIDENCE — heals never downhill; chosen lines clear "
              f"max({pol.min_recv_ros}, {pol.ros_uplift}× donor)",
              not c_bad,
              f"{len(c_bad)} lines miss their bar · HIGH {cu['HIGH']:,.0f} / "
              f"OK {cu['OK']:,.0f} / NO-DATA {cu['NO-DATA']:,.0f} u")

    # 10 · same-city split from the raw master's canonical city
    sc_bad = [x for x in lines if x["sheet"] not in ("wave2", "asm")
              and (x["sheet"] == "intra") != (bool(m_city.get(x["src"]))
                                             and m_city.get(x["src"]) == m_city.get(x["dst"]))]
    check("v_same_city",
          "SAME-CITY SPLIT — sheet 1b holds exactly the wave-1 lines whose doors "
          "share a canonical master city",
          not sc_bad,
          f"{sum(1 for x in lines if x['sheet'] == 'intra'):,} lines on 1b · "
          f"{sum(1 for x in lines if x['sheet'] == 'inter'):,} on 1 · {len(sc_bad)} misfiled")

    # 11 · master-frozen doors
    if gate("v_frozen", "Master-frozen doors on zero lines"):
        fz_l = [x for x in lines if x["src"] in frozen_m or x["dst"] in frozen_m]
        check("v_frozen",
              "FROZEN DOORS — blocked / not-Running / model-frozen (raw master) on "
              "ZERO lines", not fz_l,
              f"{len(frozen_m & plants)} frozen doors hold stock · "
              f"{len(fz_l)} lines touch one")

    # 12 · every end exists in the master
    if gate("v_in_master", "Every end in the master"):
        ghost = [x for x in lines
                 if x["src"] not in stores or x["dst"] not in stores]
        check("v_in_master",
              "CLOSED DOORS — every end exists in the store master",
              not ghost, f"{len(ghost)} lines name a door outside the master")

    # 13 · new-door donor freeze
    if gate("v_new_door", "New doors never donate"):
        nd_l = [x for x in lines if x["src"] in m_new]
        check("v_new_door",
              "NEW-DOOR DONOR FREEZE — recently opened doors ship NOTHING",
              not nd_l,
              f"{len(m_new & plants)} new doors in the live network · "
              f"{len(nd_l)} outbound lines · inbound to them "
              f"{sum(x['qty'] for x in lines if x['dst'] in m_new):,.0f} u (allowed)")

    # 14 · franchise protection — net flow + tighter transit floor
    if gate("v_fofo", "Franchise protection"):
        fofo = {sh for sh in plants
                if stores.get(sh) and stores[sh].model in pol.franchise_models} \
            - frozen_m - set(pol.afs_stores)
        nn_fofo = fofo & nn
        fofo_bad = [sh for sh in (fofo - nn_fofo) if in_u[sh] < out_u[sh] - 1e-9]
        ffl_bad = []
        for sh in fofo:
            n = norm_of(sh)
            if n > 0 and out_u[sh] > 0 and soh[sh] - out_u[sh] < \
                    min(soh[sh], pol.fofo_transit_floor * n) - 1e-6:
                ffl_bad.append(sh)
        check("v_fofo",
              f"FRANCHISE — receive ≥ give (no-norm franchise exempt) and "
              f"≥{pol.fofo_transit_floor:.0%} fill in transit, set re-derived from "
              "the master model column",
              not fofo_bad and not ffl_bad,
              f"{len(fofo)} franchise doors · {len(fofo_bad)} net-drained · "
              f"{len(ffl_bad)} below the floor · {len(nn_fofo)} exempt no-norm")

    # 15 · context norms column on the recon sheet ties to the raw context file
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    wsr = wb["5 · Store Reconciliation"]
    naw_store = Counter()
    for (sh, _c), q in naw.items():
        naw_store[sh] += q
    aw_bad, aw_rows = [], 0
    hdr, ctx_col = None, None
    for row in wsr.iter_rows(values_only=True):
        if hdr is None and row and row[0] == sid_hdr:
            hdr = list(row)
            # the operative norm column comes first, the context norm second,
            # per-category norms after — anchored on the header text, never
            # on a magic index
            norm_cols = [j for j, h in enumerate(hdr)
                         if str(h).strip().endswith(" norm")]
            assert len(norm_cols) >= 2, \
                f"recon sheet: expected operative + context norm columns, got {norm_cols}"
            ctx_col = norm_cols[1]
            continue
        sh = store_id(row[0]) if hdr is not None else None
        if sh is None or sh not in stores:
            continue
        aw_rows += 1
        v = row[ctx_col] if ctx_col < len(row) else None
        if abs(naw_store.get(sh, 0) - float(v or 0)) > 1e-6:
            aw_bad.append(sh)
    check("v_context_norms",
          "CONTEXT NORMS — the recon sheet's context-norm column ties to the raw "
          "context file, store by store",
          aw_rows > 0 and not aw_bad,
          f"{aw_rows} recon rows · {len(aw_bad)} mismatches vs the raw file")

    # 16 · style depth cap
    if gate("v_style_depth", "Style depth cap"):
        stc, sto, sti = Counter(), Counter(), Counter()
        sto1, sti1, stc1 = Counter(), Counter(), Counter()
        ssrc = defaultdict(set)
        for x in lines:
            k_in, k_out = (x["dst"], x["style"]), (x["src"], x["style"])
            sto[k_out] += x["qty"]
            sti[k_in] += x["qty"]
            if x["program"] not in ("HEAL", "HEAL3"):
                stc[k_in] += x["qty"]
            if x["sheet"] in ("inter", "intra"):
                sto1[k_out] += x["qty"]
                sti1[k_in] += x["qty"]
                if x["program"] not in ("HEAL", "HEAL3"):
                    stc1[k_in] += x["qty"]
            if x["program"] == "SET":
                ssrc[k_in].add(x["src"])
        d_bad = [k for k, q in stc1.items() if q > 0
                 and held[k] - sto1[k] + sti1[k] > max(held[k], pol.style_depth_cap) + 1e-6]
        d_bad += [k for k, q in stc.items() if q > 0
                  and held[k] - sto[k] + q > max(held[k], pol.style_depth_cap) + 1e-6]
        s_multi = {k: v for k, v in ssrc.items() if len(v) > 1}
        check("v_style_depth",
              f"STYLE DEPTH CAP — {pol.style_depth_cap} u of one style vs raw SOH; "
              "max one relocated set per door×style",
              not d_bad and (not s_multi if pol.one_set_per_door_style else True),
              f"{len(stc)} chosen style positions · {len(d_bad)} over cap · "
              f"{len(s_multi)} multi-set stacks")

    # 17 · wave 2 replay (positions replayed above)
    if gate("v_wave2", "Wave 2 replay"):
        w1_lines = [x for x in lines if x["sheet"] in ("inter", "intra")]
        w2_lines = [x for x in lines if x["sheet"] == "wave2"]
        w1_full = {(x["dst"], x["style"]) for x in w1_lines
                   if x["program"] in ("HEAL", "SET")}
        w2_pairs = {(x["dst"], x["style"]) for x in w2_lines}
        dbl = w2_pairs & w1_full
        incomplete = [k for k in w2_pairs
                      if any(pos[(k[0], k[1], z)] <= 1e-9 for z in W2)]
        donors2 = {(x["src"], x["style"]) for x in w2_lines}
        broke = [k for k in donors2
                 if all(posw1[(k[0], k[1], z)] > 1e-9 for z in W2)
                 and any(pos[(k[0], k[1], z)] <= 1e-9 for z in W2)]
        check("v_wave2",
              "WAVE 2 — every wave-2 heal lands a complete subset set; no double-serve "
              "of a wave-1 full-set receiver; no donor's existing subset set broken",
              not dbl and not incomplete and not broke,
              f"{len(w2_pairs)} wave-2 sets · {len(dbl)} double-served · "
              f"{len(incomplete)} incomplete · {len(broke)} broken donor sets")

    # 18 · assembly replay
    if gate("v_assembly", "Assembly replay"):
        asm_lines = [x for x in lines if x["sheet"] == "asm"]
        asm_pairs = {(x["dst"], x["style"]) for x in asm_lines}
        a_bad = []
        for x in asm_lines:
            if held[(x["dst"], x["style"])] > 0:
                a_bad.append(("not-virgin", x["dst"]))
            if x["sros"] > 1e-9:
                a_bad.append(("live-donor", x["src"]))
        for k in asm_pairs:
            if any(pos[(k[0], k[1], z)] <= 1e-9 for z in W2):
                a_bad.append(("incomplete", k[0]))
        per_door = Counter(d for (d, _st) in asm_pairs)
        stacked = [d for d, n2 in per_door.items() if n2 > pol.asm_per_door]
        check("v_assembly",
              "ASSEMBLY — sets land at doors with ZERO raw SOH of the style, from "
              "RoS-0 donors, complete at end, per-door cap held",
              not a_bad and not stacked,
              f"{len(asm_pairs)} assembled sets · {len(a_bad)} breaches · "
              f"{len(stacked)} doors stacked")

    # 19 · no-norm franchise exempt
    if gate("v_nnf", "No-norm franchise exempt"):
        nnf2 = {sh for sh in plants if norm_of(sh) <= 0
                and stores.get(sh) and stores[sh].model in pol.franchise_models} \
            - set(pol.afs_stores) - frozen_m
        nnf_l = [x for x in lines if x["src"] in nnf2 or x["dst"] in nnf2]
        check("v_nnf",
              "NO-NORM FRANCHISE EXEMPT — re-derived from raw norms ∩ raw master "
              "model: zero lines either way",
              not nnf_l,
              f"{len(nnf2)} doors hold {sum(soh[s] for s in nnf2):,.0f} u · "
              f"{len(nnf_l)} lines touch one")

    # 20 · fresh stock never a donor — season re-parsed from the raw holdings
    if gate("v_fresh", "Fresh stock never a donor"):
        season_age = make_season_age(pol.season_pattern, pol.season_spring_label,
                                     pol.season_current_index)
        fresh_keys = {k for k, h in holdings.items()
                      if any(season_age(s) == 0 for s, q in h.seasons.items() if q > 0)}
        fr_l = [x for x in lines if (x["src"], x["style"]) in fresh_keys]
        check("v_fresh",
              "FRESH — no line draws on a holding carrying current-season stock "
              "(season re-parsed from the raw rows)",
              not fr_l, f"{len(fresh_keys):,} fresh holdings · {len(fr_l)} violated")

    # 21 · transfer geography — re-derived from config, third proof surface
    if gate("v_geo", "Transfer geography"):
        geo_ok = make_geo_edge(stores, pol)
        g_bad = [x for x in lines if not geo_ok(x["src"], x["dst"])]
        check("v_geo",
              f"TRANSFER GEOGRAPHY — every saved line legal under "
              f"geo_mode={pol.geo_mode!r}, predicate re-derived from config",
              not g_bad, f"{len(lines):,} lines · {len(g_bad)} out of scope")

    # 22 · localization optimality — geo-constrained re-solve from the artifact
    if gate("v_localization", "Localization optimality"):
        from scipy.optimize import linear_sum_assignment
        geo_ok = make_geo_edge(stores, pol)
        tk = pol.tier_km

        def _km(a, b2):
            return pair_km(stores, a, b2, tk, aliases)

        def _legal(prog, sros, dros, a, b2):
            if a == b2 or not geo_ok(a, b2):
                return False
            if prog in ("HEAL", "HEAL3"):
                return dros >= sros - 1e-9
            if prog in ("TOPUP", "EXT"):
                return dros > sros and dros >= max(pol.min_recv_ros,
                                                   pol.ros_uplift * sros) - 1e-9
            return True

        CLASS = {"HEAL": "W1POOL", "TOPUP": "W1POOL", "EXT": "W1EXT",
                 "HEAL3": "W2", "ASM": "ASM"}
        lgroups = defaultdict(list)
        for x in lines:
            cl = CLASS.get(x["program"])
            if cl is not None:
                lgroups[(cl, x["style"], x["size"], x["qty"])].append(x)
        km_actual = km_best = 0.0
        subopt = []
        for gk, gl in lgroups.items():
            cur = sum(_km(x["src"], x["dst"]) for x in gl)
            km_actual += cur
            n = len(gl)
            if n == 1:
                km_best += cur
                continue
            C = [[(_km(gl[a]["src"], gl[b2]["dst"])
                   if _legal(gl[b2]["program"], gl[a]["sros"], gl[b2]["dros"],
                             gl[a]["src"], gl[b2]["dst"]) else 1e9)
                  for b2 in range(n)] for a in range(n)]
            rI, cI = linear_sum_assignment(C)
            best = float(sum(C[a][cI[i]] for i, a in enumerate(rI)))
            km_best += best
            if cur > best + 1e-3:
                subopt.append((gk, cur - best))
        check("v_localization",
              "DONOR LOCALIZATION — the saved pairing is the SHORTEST legal (geo-"
              "constrained) donor↔receiver matching per style×size, assignment "
              "re-solved from master coordinates + the sheet's own RoS columns",
              not subopt,
              f"{sum(len(g) for g in lgroups.values()):,} lines in "
              f"{len(lgroups):,} groups · plan {km_actual:,.0f} km vs optimum "
              f"{km_best:,.0f} km · {len(subopt)} groups improvable")

    return b
