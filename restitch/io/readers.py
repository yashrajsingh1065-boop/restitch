"""File readers + tolerant cell parsers.

One entry point — read_sheet(path, sheet) — returns a list of row-lists.
xlsb needs python-calamine (install extra: restitch[xlsb]); xlsx prefers
calamine for speed with an openpyxl fallback; csv is stdlib.
"""
from __future__ import annotations

import csv
import datetime
from pathlib import Path


class ReaderError(RuntimeError):
    pass


def _i(x):
    try:
        return int(float(x))
    except (TypeError, ValueError):
        return None


def _f(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return 0.0


def _s(x) -> str:
    return str(x).strip() if x is not None else ""


def _dt(x):
    """Tolerant date parse: real date cells arrive as datetime/date objects;
    strings and junk ('', 'NA', prose) return None."""
    if isinstance(x, datetime.datetime):
        return x
    if isinstance(x, datetime.date):
        return datetime.datetime(x.year, x.month, x.day)
    s = str(x).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


def sheet_names(path) -> list[str]:
    p = Path(path)
    suf = p.suffix.lower()
    if suf in (".xlsb", ".xlsx", ".xlsm"):
        try:
            import python_calamine as pc
            return list(pc.CalamineWorkbook.from_path(str(p)).sheet_names)
        except ImportError:
            if suf == ".xlsb":
                raise ReaderError(
                    "reading .xlsb needs python-calamine — pip install 'restitch[xlsb]'"
                ) from None
            import openpyxl
            return openpyxl.load_workbook(p, read_only=True).sheetnames
    if suf == ".csv":
        return ["csv"]
    raise ReaderError(f"unsupported file type: {p.name}")


def read_sheet(path, sheet: str | None = None) -> list[list]:
    """Full sheet as rows (lists), empty cells preserved so column indexes hold."""
    p = Path(path)
    if not p.exists():
        raise ReaderError(f"input file not found: {p}")
    suf = p.suffix.lower()
    if suf in (".xlsb", ".xlsx", ".xlsm"):
        try:
            import python_calamine as pc
            wb = pc.CalamineWorkbook.from_path(str(p))
            name = sheet or wb.sheet_names[0]
            if name not in wb.sheet_names:
                raise ReaderError(
                    f"sheet '{name}' not in {p.name}; has: {', '.join(wb.sheet_names)}")
            return wb.get_sheet_by_name(name).to_python(skip_empty_area=False)
        except ImportError:
            if suf == ".xlsb":
                raise ReaderError(
                    "reading .xlsb needs python-calamine — pip install 'restitch[xlsb]'"
                ) from None
            import openpyxl
            wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
            name = sheet or wb.sheetnames[0]
            if name not in wb.sheetnames:
                raise ReaderError(
                    f"sheet '{name}' not in {p.name}; has: "
                    f"{', '.join(wb.sheetnames)}") from None
            return [list(r) for r in wb[name].iter_rows(values_only=True)]
    if suf == ".csv":
        with open(p, newline="", encoding="utf-8-sig") as f:
            return [list(r) for r in csv.reader(f)]
    raise ReaderError(f"unsupported file type: {p.name}")
