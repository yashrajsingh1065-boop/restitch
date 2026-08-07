"""
xlsx_preflight — catch the openpyxl→Excel corruption gotchas BEFORE you ship a workbook.

Every check here maps to a real debugging session that cost ~3 hours:
  C1  leading-'=' prose cells parsed as formulas      -> feedback_openpyxl_leading_equals
  C2  empty-string cell carrying a number_format       -> feedback_openpyxl_dv_equals (#2)
  C3  DataValidation formula1 with a leading '='        -> feedback_openpyxl_dv_equals (#1)
  C4  overlapping merged-cell ranges                    -> silent data loss on open
  C5  gridlines left on                          -> no-gridlines house rule
  C6  sheet tab name >31 chars                   -> Excel truncates + "recovered records"

Usage in a builder — replace `wb.save(path)` with:

    from shared.xlsx_preflight import safe_save
    safe_save(wb, out)     # raises PreflightError on any ERROR, else saves + re-validates

Or check without saving:

    from shared.xlsx_preflight import preflight_workbook
    for issue in preflight_workbook(wb):
        print(issue)

Or validate a file someone else built (CLI):

    python3 shared/xlsx_preflight.py ~/Desktop/excels/Some_Workbook.xlsx
"""
from __future__ import annotations

import contextlib
import re
import sys
import zipfile
from dataclasses import dataclass
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import range_boundaries

ERROR = "ERROR"
WARN = "WARN"


@dataclass
class Issue:
    severity: str          # ERROR (would corrupt / break) | WARN (house-rule)
    check: str             # C1..C5
    sheet: str
    cell: str              # cell ref / range, or "" when sheet-level
    message: str
    fix: str

    def __str__(self) -> str:
        where = f"{self.sheet}!{self.cell}" if self.cell else self.sheet
        return (f"[{self.severity}] {self.check} {where}: {self.message}"
                f"\n         fix: {self.fix}")


class PreflightError(Exception):
    """Raised by safe_save when one or more ERROR-level issues are found."""

    def __init__(self, issues: list[Issue]):
        self.issues = issues
        body = "\n".join(f"  - {i}" for i in issues)
        super().__init__(f"{len(issues)} blocking issue(s) before save:\n{body}")


# --- C1: prose accidentally written as a formula -----------------------------
# openpyxl turns any string value starting with '=' into a formula cell (data_type 'f').
# A real formula has the shape  =FUNC(...)  or  =A1+B2 ;  prose like "=2PC means two-piece"
# corrupts the sheet. Heuristic: a formula body is "prose" if it contains a run of
# lowercase letters with a space and no '(' and no cell-reference pattern.
_PROSE_RE = re.compile(r"[a-z]{2,}\s+[a-z]{2,}")          # two+ lowercase words in a row
_CELLREF_RE = re.compile(r"[A-Z]{1,3}\$?\d+|\b[A-Z]+\s*\(")  # A1 / $B$2 / SUM(


def _looks_like_prose_formula(formula: str) -> bool:
    body = formula[1:] if formula.startswith("=") else formula
    if "(" in body or _CELLREF_RE.search(body):
        return False                     # has a func call or cell ref -> real formula
    return bool(_PROSE_RE.search(body))


# --- the scanner --------------------------------------------------------------
def preflight_workbook(wb: Workbook) -> list[Issue]:
    """Scan an in-memory Workbook for the known Excel-corruption gotchas."""
    issues: list[Issue] = []

    for ws in wb.worksheets:
        # C6 — tab name over Excel's hard 31-char limit. openpyxl only WARNS and keeps the
        # long name; Excel then truncates/renames the sheet on open and shows the dreaded
        # "we found a problem… recovered some data" prompt. Treat as blocking.
        if len(ws.title) > 31:
            issues.append(Issue(
                ERROR, "C6", ws.title, "",
                f"sheet tab name is {len(ws.title)} chars (>31) — Excel "
                "truncates it and reports recovered records on open",
                "shorten the sheet title to 31 characters or fewer",
            ))

        # C5 — gridlines (house rule: every sheet showGridLines = False).
        # openpyxl's default is None, which Excel renders as gridlines-ON — so anything
        # that isn't an explicit False counts as "on".
        if ws.sheet_view.showGridLines is not False:
            issues.append(Issue(
                WARN, "C5", ws.title, "",
                "gridlines are ON",
                "ws.sheet_view.showGridLines = False",
            ))

        # C3 — DataValidation formula1 with leading '='
        for dv in ws.data_validations.dataValidation:
            f1 = dv.formula1
            if isinstance(f1, str) and f1.startswith("="):
                cells = str(dv.sqref) if dv.sqref else ""
                issues.append(Issue(
                    ERROR, "C3", ws.title, cells,
                    f"DataValidation formula1 has a leading '=' ({f1!r})",
                    "drop the '=' — list/range validations want "
                    "'\"a,b,c\"' or 'rng' without '='",
                ))

        # C4 — overlapping merged ranges
        merges = [str(r) for r in ws.merged_cells.ranges]
        for a, b in _overlapping_pairs(merges):
            issues.append(Issue(
                ERROR, "C4", ws.title, f"{a} & {b}",
                "merged ranges overlap (Excel drops the whole merge block on open)",
                "make the ranges disjoint; never merge a cell that's already in another merge",
            ))

        # C1 + C2 — per-cell scans
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                # C1 prose-as-formula
                if cell.data_type == "f" and isinstance(val, str) \
                        and _looks_like_prose_formula(val):
                    issues.append(Issue(
                        ERROR, "C1", ws.title, cell.coordinate,
                        f"text starting with '=' was stored as a formula ({val!r})",
                        "prefix the text with an apostrophe, or "
                        "rephrase so it doesn't start with '='",
                    ))
                # C2 empty string carrying a number_format
                if val == "" and cell.number_format not in ("General", None):
                    issues.append(Issue(
                        ERROR, "C2", ws.title, cell.coordinate,
                        f"empty string with number_format="
                        f"{cell.number_format!r} (writes malformed inlineStr)",
                        "set the cell to None instead of '' (leave it "
                        "blank), or reset number_format to General",
                    ))

    return issues


def _overlapping_pairs(ranges: list[str]):
    boxes = []
    for r in ranges:
        min_c, min_r, max_c, max_r = range_boundaries(r)
        boxes.append((r, min_c, min_r, max_c, max_r))
    for i in range(len(boxes)):
        for j in range(i + 1, len(boxes)):
            ra, c1, r1, c2, r2 = boxes[i]
            rb, c3, r3, c4, r4 = boxes[j]
            if c1 <= c4 and c3 <= c2 and r1 <= r4 and r3 <= r2:
                yield ra, rb


# --- post-write validation ----------------------------------------------------
def validate_xlsx_file(path: str | Path) -> list[Issue]:
    """Open a written .xlsx the way Excel would, and confirm its XML is well-formed.
    Catches gross corruption that only shows up once the file is zipped to disk."""
    path = Path(path)
    issues: list[Issue] = []

    # 1) every part must be well-formed XML
    try:
        with zipfile.ZipFile(path) as z:
            for name in z.namelist():
                if name.endswith(".xml") or name.endswith(".rels"):
                    try:
                        ET.fromstring(z.read(name))
                    except ET.ParseError as e:
                        issues.append(Issue(
                            ERROR, "FILE", name, "",
                            f"malformed XML part: {e}",
                            "rebuild the workbook; a cell/validation wrote invalid XML",
                        ))
    except zipfile.BadZipFile:
        issues.append(Issue(ERROR, "FILE", str(path), "",
                            "not a valid .xlsx (zip) file", "rebuild"))
        return issues

    # 2) it must re-open cleanly in openpyxl (proxy for "Excel can load it")
    try:
        load_workbook(path)
    except Exception as e:  # noqa: BLE001 — any load failure is a real defect
        issues.append(Issue(
            ERROR, "FILE", str(path), "",
            f"openpyxl cannot reopen the file: {type(e).__name__}: {e}",
            "rebuild the workbook",
        ))

    return issues


# --- the one-liner builders should call --------------------------------------
def safe_save(wb: Workbook, path: str | Path, *, strict_gridlines: bool = False) -> list[Issue]:
    """Pre-flight, save, then re-validate the file on disk.

    Raises PreflightError if any ERROR-level issue is found (before OR after save).
    WARN issues (e.g. gridlines on) are returned, not raised — unless strict_gridlines=True.
    Returns the list of WARN issues so the caller can log them.
    """
    pre = preflight_workbook(wb)
    blocking = [i for i in pre if i.severity == ERROR]
    if strict_gridlines:
        blocking += [i for i in pre if i.check == "C5"]
    if blocking:
        raise PreflightError(blocking)

    wb.save(path)

    post = validate_xlsx_file(path)
    if post:
        raise PreflightError(post)

    return [i for i in pre if i.severity == WARN]


# --- CLI: validate an existing file ------------------------------------------
def _cli(argv: list[str]) -> int:
    if not argv:
        print("usage: python3 xlsx_preflight.py <file.xlsx> [more.xlsx ...]")
        return 2
    rc = 0
    for p in argv:
        issues = validate_xlsx_file(p)
        # also re-scan the loaded workbook for the in-memory checks
        with contextlib.suppress(Exception):
            issues = preflight_workbook(load_workbook(p)) + issues
        errors = [i for i in issues if i.severity == ERROR]
        warns = [i for i in issues if i.severity == WARN]
        status = "FAIL" if errors else ("WARN" if warns else "OK")
        print(f"\n{p}  ->  {status}  ({len(errors)} error, {len(warns)} warn)")
        for i in issues:
            print(f"  {i}")
        if errors:
            rc = 1
    return rc


if __name__ == "__main__":
    raise SystemExit(_cli(sys.argv[1:]))
