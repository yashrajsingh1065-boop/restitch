"""House workbook style — the formatting vocabulary every sheet shares.

Extracted from the source project's builder so the dead legacy module can be
retired. Conventions: no gridlines, slate/navy headers, zebra bands, KPI tiles,
prose callouts. One vocabulary across every sheet — consistency IS an
affordance in a workbook someone must trust before loading a truck.
"""
from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ── palette ──
NAVY = "1F2A44"
SLATE = "334155"
STEEL = "475569"
LSLATE = "F1F5F9"
LSTEEL = "E2E8F0"
BAND = "F8FAFC"
TEAL = "0F766E"
MAROON = "9F1239"
GREEN = "15803D"
AMBER = "B45309"
WHITE = "FFFFFF"
INK = "0F172A"
MUTE = "64748B"
GOLD = "F5F3E7"

# ── number formats ──
INT = "#,##0"
PCT = "0.0%"
ROS2 = "0.00"


def F(sz=10, b=False, color=INK, italic=False):
    return Font(name="Calibri", size=sz, bold=b, color=color, italic=italic)


def fill(c):
    return PatternFill("solid", fgColor=c)


def side(c="CBD5E1", st="thin"):
    return Side(style=st, color=c)


def border(c="CBD5E1", st="thin"):
    return Border(side(c, st), side(c, st), side(c, st), side(c, st))


LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
RGT = Alignment(horizontal="right", vertical="center")
TL = Alignment(horizontal="left", vertical="top", wrap_text=True)


def nogrid(ws):
    ws.sheet_view.showGridLines = False
    return ws


def widths(ws, ws_widths):
    for col, w in ws_widths.items():
        ws.column_dimensions[col].width = w


def sheet_title(index: int, name: str, sep: str = " · ") -> str:
    """Tab name with the 31-char Excel limit enforced — openpyxl only warns,
    Excel then truncates and reports 'recovered records'."""
    t = f"{index}{sep}{name}"
    return t[:31]


def title_block(ws, title, subtitle, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, title)
    c.font = F(17, True, WHITE)
    c.fill = fill(NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(2, 1, subtitle)
    c.font = F(9, False, WHITE, True)
    c.fill = fill(SLATE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 16


def header_row(ws, r, headers, c0=1, fillc=SLATE, fontc=WHITE, sz=9):
    for j, h in enumerate(headers):
        cell = ws.cell(r, c0 + j, h)
        cell.font = F(sz, True, fontc)
        cell.fill = fill(fillc)
        cell.alignment = CTR
        cell.border = border("94A3B8")
    ws.row_dimensions[r].height = 26


def datarow(ws, r, vals, c0=1, band=False, aligns=None, fmts=None, fontc=INK, bold=False):
    for j, v in enumerate(vals):
        cell = ws.cell(r, c0 + j, v)
        cell.font = F(9, bold, fontc)
        cell.alignment = (aligns[j] if aligns else LEFT)
        cell.border = border()
        if band:
            cell.fill = fill(BAND)
        if fmts and fmts[j]:
            cell.number_format = fmts[j]


def kpi(ws, r, c, label, value, sub="", vcolor=NAVY, w=2):
    ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + w - 1)
    cell = ws.cell(r, c, label)
    cell.font = F(8.5, True, MUTE)
    cell.fill = fill(LSLATE)
    cell.alignment = Alignment("left", indent=1)
    ws.merge_cells(start_row=r + 1, start_column=c, end_row=r + 1, end_column=c + w - 1)
    cell = ws.cell(r + 1, c, value)
    cell.font = F(19, True, vcolor)
    cell.fill = fill(LSLATE)
    cell.alignment = Alignment("left", indent=1)
    ws.merge_cells(start_row=r + 2, start_column=c, end_row=r + 2, end_column=c + w - 1)
    cell = ws.cell(r + 2, c, sub)
    cell.font = F(8, False, STEEL)
    cell.fill = fill(LSLATE)
    cell.alignment = Alignment("left", indent=1)
    for rr in (r, r + 1, r + 2):
        ws.cell(rr, c).border = Border(top=side("CBD5E1"), bottom=side("CBD5E1"),
                                       left=side("CBD5E1"), right=side("CBD5E1"))
    ws.row_dimensions[r].height = 15
    ws.row_dimensions[r + 1].height = 24
    ws.row_dimensions[r + 2].height = 13


def callout(ws, r, ncols, text, kind="info"):
    bg = {"info": LSLATE, "warn": GOLD, "good": "ECFDF5", "bad": "FEF2F2"}[kind]
    fg = {"info": SLATE, "warn": AMBER, "good": GREEN, "bad": MAROON}[kind]
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(r, 1, text)
    c.font = F(9.5, False, fg)
    c.fill = fill(bg)
    c.alignment = Alignment("left", "center", wrap_text=True, indent=1)
    c.border = Border(left=side(fg, "medium"))
    ws.row_dimensions[r].height = 30
