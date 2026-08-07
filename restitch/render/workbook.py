"""The 13-sheet movement workbook — every label generated from config.

Ported from the source project's renderer and de-tenanted: category columns
come from the run's category list, wave-2 labels from the policy's subset,
warehouse columns from the declared warehouse codes, period tags from the
`labels` dict a tenant passes in. The layout is the layout that survived a
buyer's daily use: movement sheets first, the working shown, checks with
evidence, provenance in every subtitle.

The battery (core.checks.run_battery) runs BEFORE anything renders and a
failed check refuses to save — this file instructs physical stock movement
that cannot be undone, so a workbook that exists is a workbook that passed.
"""
from __future__ import annotations

from collections import Counter, defaultdict

import openpyxl

from ..core.checks import build_recon, run_battery
from ..core.ids import id_sort_key
from ..core.model import canon_city, haversine
from ..core.policy import PRIO, PROGRAMS
from ..core.rules import RULES
from .preflight import safe_save
from .style import (
    AMBER,
    CTR,
    GREEN,
    INK,
    INT,
    LEFT,
    MAROON,
    MUTE,
    NAVY,
    PCT,
    RGT,
    ROS2,
    SLATE,
    TEAL,
    F,
    border,
    callout,
    datarow,
    fill,
    header_row,
    kpi,
    nogrid,
    sheet_title,
    title_block,
    widths,
)

BAND_RANK = {"P1": 0, "P2": 1, "P3": 2}
BAND_COL = {"P1": GREEN, "P2": AMBER, "P3": MUTE}
CONF_COL = {"HIGH": GREEN, "OK": AMBER, "NO-DATA": MUTE, "DEPT": TEAL}

DEFAULT_LABELS = dict(
    title="Cut-size redeployment — direct store-to-store movement file",
    currency="",           # e.g. "₹" — appended to money headers
    soh_asof="",           # display string for the stock as-of date
    norm="norm",           # operative norm period tag, e.g. "SS26"
    norm_context="context",  # context norm period tag, e.g. "AW26"
    tier_label="tier",     # prefix for tier row labels on the fill sheet
    ros_note="",           # provenance prose for the velocity source
    fs_note="",            # provenance prose for the free-stock file
    evidence_prefix="",    # velocity evidence field prefix, e.g. "d6_"
    store_id="Store ID",   # store-id column header (a tenant may pass its own vocabulary)
)


def idc(v):
    """Store-id CELL value: canonical ids are str; all-digit ids write as
    numbers so numeric-code tenants keep clean cells (and the legacy render
    parity gate keeps comparing equal numeric cells). Alpha ids verbatim."""
    return int(v) if isinstance(v, str) and v.isdigit() else v


def _prog_label(pol):
    w2 = "·".join(pol.wave2_pivotals)
    return {"HEAL": "completes a cut set", "SET": "whole dead set relocation",
            "HEAL3": f"wave 2 — completes {w2} set",
            "ASM": f"assembled {w2} set at dept-selling door",
            "TOPUP": "depth into higher-RoS full set", "EXT": "extreme size to seller"}


def enrich_rows(R, pol, labels):
    """Stamp each move with the same-city flag and the receiver's velocity
    evidence — display attributes the sheets and battery partition on."""
    aliases = dict(pol.city_aliases)
    city = {sh: canon_city(s.city, aliases) for sh, s in R["stores"].items()}
    pre = labels["evidence_prefix"]
    for x in R["moves"]:
        x["same_city"] = bool(city.get(x["src"])) and city.get(x["src"]) == city.get(x["dst"])
        rec = R["perf"].get((x["dst"], x["style"])) or {}
        x["d_sold8"] = rec.get(f"{pre}sold8", 0.0) if pre else 0.0
        x["d_grn8"] = rec.get(f"{pre}grn8", 0.0) if pre else 0.0
    return R["moves"]


def _metrics(R, rows):
    fl = R["fill"]

    def _p(s):
        return s["cut"]["ALL"] / s["tot"]["ALL"] if s["tot"]["ALL"] else 0

    cut_h = [h for h in R["holdings"].values() if h.cut]
    return dict(
        cut_u=sum(h.units for h in cut_h), cut_h=len(cut_h),
        cut_pct_u=sum(h.units for h in cut_h) / R["opening_units"],
        cut_p1=_p(fl["p1"]), cut_full=_p(fl["full"]),
        resolved=len(R["completed"]) + len(R["healed_by_set"]),
        gap_struct=len(R["gap_structural"]), gap_prio=len(R["gap_priority"]),
        moved_u=sum(x["qty"] for x in rows),
        p1_units=sum(x["qty"] for x in rows if x["band"] == "P1"),
        p1_lines=sum(1 for x in rows if x["band"] == "P1"),
        uphill_u=sum(x["qty"] for x in rows if x["dst_ros"] > x["src_ros"]),
        reloc_n=len(R["relocations"]), core_u=R["core_stat"]["units"],
        gap_cap=len(R["gap_capacity"]),
        conf_hi=sum(x["qty"] for x in rows if x.get("conf") == "HIGH"))


# ───────────────────────── shared movement table ─────────────────────────
def _moves_table(ws, R, pol, sel, r, cur, samecity_col=False):
    from openpyxl.utils import get_column_letter as L
    stores = R["stores"]
    hdr = ["Priority", "Program", "Take from", "From store", "From city", "From state",
           "From rgn", "From type", "From L2L", "From format", "→",
           "Deliver to", "To store", "To city", "To state", "To rgn", "To type", "To L2L",
           "To format", "Style", "Cat", "Size", "Qty", f"MRP {cur}".strip(),
           f"Value {cur}".strip(), "RoS from", "RoS to", "Recv sold 8wk",
           "Recv GRN 8wk", "Confidence"]
    if samecity_col:
        hdr.append("Same city")
    NC = len(hdr)
    header_row(ws, r, hdr)
    hr = r
    r += 1
    al = ([CTR, CTR, CTR, LEFT, LEFT, LEFT, CTR, CTR, CTR, LEFT, CTR,
           CTR, LEFT, LEFT, LEFT, CTR, CTR, CTR, LEFT, LEFT, CTR, CTR,
           RGT, RGT, RGT, RGT, RGT, RGT, RGT, CTR] + ([CTR] if samecity_col else []))
    fm = ([None] * 22 + [INT, INT, INT, ROS2, ROS2, INT, INT, None]
          + ([None] if samecity_col else []))
    for i, x in enumerate(sorted(sel, key=lambda z: (BAND_RANK[z["band"]],
                                                     PRIO[z["program"]],
                                                     -z["value"], str(z["style"]),
                                                     z["size"]))):
        ss, ds = stores[x["src"]], stores[x["dst"]]
        vals = [x["band"], x["program"], idc(x["src"]), ss.name, ss.city, ss.state, ss.region,
                ss.stype, ss.l2l, ss.fmt, "→",
                idc(x["dst"]), ds.name, ds.city, ds.state, ds.region, ds.stype, ds.l2l, ds.fmt,
                x["style"], x["cat"], x["size"],
                x["qty"], x["mrp"], x["value"], x["src_ros"], x["dst_ros"],
                x.get("d_sold8", 0), x.get("d_grn8", 0), x.get("conf", "")]
        if samecity_col:
            vals.append("YES" if x.get("same_city") else "no")
        datarow(ws, r, vals, band=(i % 2 == 0), aligns=al, fmts=fm)
        ws.cell(r, 1).font = F(9, True, BAND_COL[x["band"]])
        ws.cell(r, 30).font = F(8, True, CONF_COL.get(x.get("conf", ""), INK))
        r += 1
    ws.auto_filter.ref = f"A{hr}:{L(NC)}{r - 1}"
    widths(ws, {"A": 9, "B": 9, "C": 11, "D": 28, "E": 13, "F": 13, "G": 8, "H": 11,
                "I": 7, "J": 16, "K": 4, "L": 10, "M": 28, "N": 13, "O": 13, "P": 8,
                "Q": 11, "R": 7, "S": 16, "T": 16, "U": 8, "V": 6, "W": 7, "X": 9,
                "Y": 11, "Z": 9, "AA": 9, "AB": 11, "AC": 11, "AD": 11, "AE": 9})
    return r


# ───────────────────────── sheets ─────────────────────────
def sheet_moves(wb, R, pol, rows, lb):
    w1 = [x for x in rows if x["program"] not in ("HEAL3", "ASM")]
    inter = [x for x in w1 if not x.get("same_city")]
    intra_n = len(w1) - len(inter)
    ws = nogrid(wb.create_sheet("1 · Movement — Take & Deliver"))
    NC = 20
    tot = sum(x["qty"] for x in inter)
    ls = R["loc_stat"]
    loc_note = (f" · donors localized "
                f"(−{(1 - ls['km_after'] / ls['km_before']) * 100:.0f}% distance)"
                if ls["km_before"] else "")
    title_block(ws, "TAKE FROM → DELIVER TO — every INTER-CITY unit, direct store to store",
                f"{len(inter):,} lines · {tot:,.0f} units · SOH {lb['soh_asof']}"
                f"{' · ' + lb['ros_note'] if lb['ros_note'] else ''}"
                f" · same-city transfers on sheet 1b{loc_note}", NC)
    r = 4
    bu, bn = Counter(), Counter()
    for x in inter:
        bu[x["band"]] += x["qty"]
        bn[x["band"]] += 1
    for i, bnd in enumerate(["P1", "P2", "P3"]):
        kpi(ws, r, 1 + i * 2,
            f"{bnd} — "
            f"{'execute now' if bnd == 'P1' else ('next tranche' if bnd == 'P2' else 'defer')}",
            f"{bu[bnd]:,.0f} u", f"{bn[bnd]:,} lines", BAND_COL[bnd])
    up = sum(x["qty"] for x in inter if x["dst_ros"] > x["src_ros"])
    kpi(ws, r, 7, "RoS-uphill units", f"{up / tot:.0%}" if tot else "—",
        f"{up:,.0f} u to a better-RoS door", NAVY)
    hi = sum(x["qty"] for x in inter if x.get("conf") == "HIGH")
    kpi(ws, r, 9, "HIGH-confidence units", f"{hi / tot:.0%}" if tot else "—",
        f"receiver RoS ≥ {pol.min_recv_ros:.2f} and ≥ {pol.ros_uplift:.1f}× donor", TEAL)
    r += 4
    callout(ws, r, NC,
            "EVERY LINE IS ACTIONABLE AS WRITTEN — one direct transfer, donor door to "
            "receiver door. No warehouse legs, no holds: every id on this sheet is a "
            "store that holds category stock today. 'RoS from/to' is THAT style's rate "
            "at that door; Confidence grades the pair (HIGH = real seller with a clear "
            f"uplift · OK = uphill but modest · NO-DATA = no velocity either end). "
            f"Same-city transfers sit on sheet 1b ({intra_n:,} lines) — execute as "
            f"local runs. Filter Priority = P1 to execute the "
            f"{pol.movement_budget:,}-unit plan.", "info")
    r += 4
    _moves_table(ws, R, pol, inter, r, lb["currency"])
    return ws


def sheet_samecity(wb, R, pol, rows, lb):
    intra = [x for x in rows if x.get("same_city") and x["program"] not in ("HEAL3", "ASM")]
    ws = nogrid(wb.create_sheet("1b · Same-City Moves"))
    NC = 20
    tot = sum(x["qty"] for x in intra)
    aliases = dict(pol.city_aliases)
    cities = sorted({canon_city(R["stores"][x["src"]].city, aliases) for x in intra})
    title_block(ws, "SAME-CITY TRANSFERS — donor and receiver in the same city",
                f"{len(intra):,} lines · {tot:,.0f} units · {len(cities)} cities · "
                f"split out for local execution · SOH {lb['soh_asof']}", NC)
    r = 4
    bu = Counter()
    for x in intra:
        bu[x["band"]] += x["qty"]
    kpi(ws, r, 1, "Same-city units", f"{tot:,.0f}", f"{len(intra):,} lines", NAVY)
    kpi(ws, r, 3, "P1 — execute now", f"{bu['P1']:,.0f} u", "same budget, same checks", GREEN)
    kpi(ws, r, 5, "Cities involved", f"{len(cities)}", "local van runs", TEAL)
    hi = sum(x["qty"] for x in intra if x.get("conf") == "HIGH")
    kpi(ws, r, 7, "HIGH-confidence units", f"{hi / tot:.0%}" if tot else "—",
        f"{hi:,.0f} u", GREEN)
    r += 4
    callout(ws, r, NC,
            "These lines are IDENTICAL in status to sheet 1 — same plan, same budget, "
            "same check battery — separated only because both doors sit in the same "
            "city, so they execute as local runs rather than line-haul. Every "
            "eligibility rule applies exactly as on sheet 1.", "info")
    r += 4
    _moves_table(ws, R, pol, intra, r, lb["currency"])
    return ws


def sheet_wave2(wb, R, pol, rows, lb):
    w2lbl = "·".join(pol.wave2_pivotals)
    w2r = [x for x in rows if x["program"] == "HEAL3"]
    w2 = R["wave2_stat"]
    ws = nogrid(wb.create_sheet(sheet_title("1c", f"Wave 2 — {w2lbl}")))
    NC = 20
    tot = sum(x["qty"] for x in w2r)
    title_block(ws, f"WAVE 2 — {w2lbl} set completion on the plan's AFTER position",
                f"{len(w2r):,} lines · {tot:,.0f} units · {w2['sets']:,} new subset "
                f"sets · executes AFTER sheets 1/1b have landed · SOH {lb['soh_asof']}", NC)
    r = 4
    (c3o, t3o), (c3w, t3w), (c3e, t3e) = w2["cut3_open"], w2["cut3_w1"], w2["cut3_end"]
    kpi(ws, r, 1, "New subset sets created", f"{w2['sets']:,}", f"{tot:,.0f} u", GREEN)
    kpi(ws, r, 3, f"{w2lbl} gap — after wave 1",
        f"{c3w / t3w:.1%}" if t3w else "—", f"{c3w:,.0f} u sit in subset-cut holdings", MAROON)
    kpi(ws, r, 5, f"{w2lbl} gap — after wave 2",
        f"{c3e / t3e:.1%}" if t3e else "—",
        f"was {c3o / t3o:.1%} at opening" if t3o else "", NAVY)
    kpi(ws, r, 7, "Blocked by the peak cap", f"{w2['cap_blocked']:,}",
        "doors already full from wave-1 chosen inbound", AMBER)
    r += 4
    callout(ws, r, NC,
            "WHAT THIS SHEET IS: take the movement plan's AFTER stock position as the "
            f"new opening, and complete sets under the {w2lbl} subset lens. EXECUTE "
            "ONLY AFTER SHEETS 1/1b HAVE LANDED — its donor positions assume wave 1 "
            "has moved. Every constraint of the main plan carries over unchanged. "
            "Three wave-2-specific rules: a door that received a full set of a style "
            "in wave 1 never receives that style again · a donation never breaks an "
            "existing subset set at the donor · every wave-1 cap guarantee survives.",
            "info")
    r += 4
    _moves_table(ws, R, pol, w2r, r, lb["currency"], samecity_col=True)
    return ws


def sheet_assembly(wb, R, pol, rows, lb):
    w2lbl = "·".join(pol.wave2_pivotals)
    ar = [x for x in rows if x["program"] == "ASM"]
    am = R["asm_stat"]
    ws = nogrid(wb.create_sheet("1d · Assembly — New Sets"))
    NC = 20
    tot = sum(x["qty"] for x in ar)
    title_block(ws, f"ASSEMBLY — new {w2lbl} sets built at department-selling doors",
                f"{len(ar):,} lines · {tot:,.0f} units · {am['sets']:,} sets at "
                f"{am['doors']:,} doors · executes with wave 2, AFTER sheets 1/1b · "
                f"SOH {lb['soh_asof']}", NC)
    r = 4
    kpi(ws, r, 1, "New sets assembled", f"{am['sets']:,}", f"{am['styles']:,} styles", GREEN)
    kpi(ws, r, 3, "Units", f"{tot:,.0f}", "dead donor stock only", NAVY)
    kpi(ws, r, 5, "Doors gaining a NEW style", f"{am['doors']:,}",
        f"max {pol.asm_per_door} assembled set(s) per door", TEAL)
    kpi(ws, r, 7, "Donor RoS", "0.00", "every source is DEAD stock — pure uphill", MAROON)
    r += 4
    callout(ws, r, NC,
            "Sets are built ONLY from DEAD leftover pool stock, ONLY at doors that "
            "already SELL the style's DEPARTMENT (ranked by department sales, best "
            "first), capped per door. Every chosen-program gate binds in full. "
            "Confidence reads DEPT — the destination is picked on department velocity "
            "because a style-RoS ladder cannot exist at a door that never carried "
            "the style.", "info")
    r += 4
    _moves_table(ws, R, pol, ar, r, lb["currency"], samecity_col=True)
    return ws


def sheet_fill(wb, R, pol, m, lb):
    fl = R["fill"]
    ws = nogrid(wb.create_sheet("2 · Fill Rate — Before & After"))
    NC = 8
    title_block(ws, "Fill rate & cut% — before, during, after",
                f"'during' = the {pol.movement_budget:,}-unit P1 plan executed · "
                f"SOH {lb['soh_asof']}", NC)
    r = 4
    kpi(ws, r, 1, "Cut units — before", f"{m['cut_u']:,.0f}",
        f"{m['cut_pct_u']:.1%} of stock", MAROON)
    kpi(ws, r, 3, "Cut sets resolved", f"{m['resolved']:,}", "full plan", NAVY)
    kpi(ws, r, 5, "Units moved — full plan", f"{m['moved_u']:,.0f}",
        f"{m['p1_units']:,.0f} funded at P1", GREEN)
    kpi(ws, r, 7, "Fresh-buy gaps", f"{m['gap_struct']:,}", "size absent NATIONALLY", AMBER)
    r += 4
    tiers_ = sorted({t for t in fl["before"]["tot"] if t != "ALL"})
    ndoor = Counter()
    for sh in {s for (s, _st) in R["holdings"]}:
        ndoor[R["tiers"].get(sh, "—")] += 1
        ndoor["ALL"] += 1
    states = [("BEFORE — today's SOH", fl["before"]),
              (f"DURING — after P1 ({pol.movement_budget / 1000:,.0f}k u)", fl["p1"]),
              ("AFTER — full plan", fl["full"])]
    cats = " + ".join(R["categories"])
    for title, key in (("CUT-SET % — store units sitting in a size-broken set", "cut"),
                       (f"FILL RATE vs {lb['norm']} {cats} norm", None)):
        ws.cell(r, 1, title).font = F(11, True, NAVY)
        r += 1
        header_row(ws, r, ["Scope", "Doors", "Base u"] + [s[0] for s in states]
                   + ["Before → After"])
        r += 1
        for i, (lab, k) in enumerate([("NETWORK — all doors", "ALL")]
                                     + [(f"{lb['tier_label']} {t}", t) for t in tiers_]):
            vals = []
            for _, s in states:
                if key:
                    vals.append(s["cut"][k] / s["tot"][k] if s["tot"][k] else 0.0)
                else:
                    vals.append(s["soh"][k] / s["norm"][k] if s["norm"][k] else 0.0)
            thin = ndoor[k] <= 3
            datarow(ws, r, [lab + ("  ⚠ thin base" if thin else ""), ndoor[k],
                            fl["before"]["tot"][k]] + vals + [vals[-1] - vals[0]],
                    band=(i % 2 == 0), aligns=[LEFT, RGT, RGT, RGT, RGT, RGT, RGT],
                    fmts=[None, INT, INT, PCT, PCT, PCT, PCT],
                    fontc=(MUTE if thin else INK), bold=(k == "ALL"))
            r += 1
        r += 1

    def _p(s):
        return s["cut"]["ALL"] / s["tot"]["ALL"] if s["tot"]["ALL"] else 0

    callout(ws, r, NC,
            f"Cut-set % falls {_p(fl['before']):.0%} → {_p(fl['p1']):.0%} on the "
            f"{pol.movement_budget:,}-unit P1 plan and {_p(fl['full']):.0%} if "
            "everything runs. Network units DO NOT leave the stores — every move lands "
            "at another door, so fill moves by redistribution only. Thin-base tiers "
            "are greyed — they swing on a handful of units.", "info")
    r += 4
    widths(ws, {"A": 46, "B": 9, "C": 11, "D": 22, "E": 22, "F": 20, "G": 16, "H": 12})
    return ws


def sheet_soh(wb, R, pol, recon, idx, lb):
    from openpyxl.utils import get_column_letter as L
    SZ = list(pol.size_order)
    PIVOTALS = list(pol.pivotals)
    is_core = pol.is_core
    stores, H, tiers_, post = R["stores"], R["holdings"], R["tiers"], recon["post"]
    created = [k for k, sz in post.items()
               if k not in H and any(q > 1e-9 for q in sz.values())]
    stylemeta = {}
    for (_sh, _st), h in H.items():
        stylemeta.setdefault(_st, (h.div, h.cat))
    keys = sorted(list(H) + created,
                  key=lambda k: (id_sort_key(k[0]), k[1]))
    name = f"{idx} · SOH — Before & After"
    ws = nogrid(wb.create_sheet(name))
    SZSET = set(SZ)
    W = len(SZ) + 1
    c_b0 = 15
    c_a0 = c_b0 + W
    c_ub, c_ua, c_cb, c_ca = c_a0 + W, c_a0 + W + 1, c_a0 + W + 2, c_a0 + W + 3
    NC = c_ca
    oth = sorted({z for h in H.values() for z in h.sizes if z not in SZSET})
    title_block(ws, "SOH — every store × style × size, before and after the plan",
                f"{len(keys):,} rows ({len(created):,} holdings created by the plan) · "
                "AFTER = opening stock with every movement line applied", NC)
    last = 10 + len(keys)
    for i, (lab, form, fmt, col) in enumerate([
            ("Units BEFORE", f"=SUM({L(c_ub)}11:{L(c_ub)}{last})", INT, NAVY),
            ("Units AFTER", f"=SUM({L(c_ua)}11:{L(c_ua)}{last})", INT, NAVY),
            ("Cut units BEFORE",
             f"=SUMIF({L(c_cb)}11:{L(c_cb)}{last},\"CUT\",{L(c_ub)}11:{L(c_ub)}{last})",
             INT, MAROON),
            ("Cut units AFTER",
             f"=SUMIF({L(c_ca)}11:{L(c_ca)}{last},\"CUT\",{L(c_ua)}11:{L(c_ua)}{last})",
             INT, GREEN)]):
        ws.cell(5 + i, 2, lab).font = F(9, True, SLATE)
        c = ws.cell(5 + i, 3, form)
        c.number_format = fmt
        c.font = F(11, True, col)
        c.alignment = RGT
    callout(ws, 9, NC,
            "The four cells above are live formulas over this sheet. Cut% BEFORE = "
            "C7/C5, cut% AFTER = C8/C6 — the Cut% Working sheet shows that derivation. "
            "'b' columns are stock today, 'a' the same holding once every movement "
            "line is applied. Units BEFORE equals units AFTER at network level: direct "
            "moves never take stock out of the store network."
            + (f" 'oth' aggregates sizes outside the ladder ({', '.join(oth)})."
               if oth else "")
            + " FROZEN? flags styles frozen by policy — their rows are identical "
              "before vs after.", "info")
    hdr = ([lb["store_id"], "Store", "City", "Region", "State", "Type",
            "Model", "L2L", "Format",
            "Tier", "Div", "Cat", "Style", "FROZEN?"]
           + [f"{s} b" for s in SZ] + ["oth b"] + [f"{s} a" for s in SZ] + ["oth a"]
           + ["Units before", "Units after", "Cut before", "Cut after"])
    header_row(ws, 10, hdr)
    for k, sz in enumerate(SZ):
        if sz in PIVOTALS:
            ws.cell(10, c_b0 + k).fill = fill(TEAL)
            ws.cell(10, c_a0 + k).fill = fill(TEAL)
    pivB = [L(c_b0 + SZ.index(p)) for p in PIVOTALS]
    pivA = [L(c_a0 + SZ.index(p)) for p in PIVOTALS]
    r = 11
    disp_b = disp_a = 0.0
    for (shrm, style) in keys:
        h = H.get((shrm, style))
        st = stores.get(shrm)
        div, cat = (h.div, h.cat) if h else stylemeta.get(style, ("", ""))
        for c, v in ((1, idc(shrm)), (2, st.name if st else ""), (3, st.city if st else ""),
                     (4, st.region if st else ""), (5, st.state if st else ""),
                     (6, st.stype if st else ""), (7, st.model if st else ""),
                     (8, st.l2l if st else ""), (9, st.fmt if st else ""),
                     (10, tiers_.get(shrm, "")), (11, div), (12, cat), (13, style),
                     (14, "FROZEN" if is_core(style) else "")):
            ws.cell(r, c, v)
        aft = post.get((shrm, style), {})
        for k, sz in enumerate(SZ):
            qb = h.sizes.get(sz) if h else None
            if qb:
                ws.cell(r, c_b0 + k, qb)
                disp_b += qb
            qa = aft.get(sz)
            if qa and qa > 1e-9:
                ws.cell(r, c_a0 + k, qa)
                disp_a += qa
        ob = sum(q for sz, q in (h.sizes.items() if h else ()) if sz not in SZSET and q)
        if ob:
            ws.cell(r, c_b0 + len(SZ), ob)
            disp_b += ob
        oa = sum(q for sz, q in aft.items() if sz not in SZSET and q > 1e-9)
        if oa:
            ws.cell(r, c_a0 + len(SZ), oa)
            disp_a += oa
        ws.cell(r, c_ub, f"=SUM({L(c_b0)}{r}:{L(c_b0 + W - 1)}{r})")
        ws.cell(r, c_ua, f"=SUM({L(c_a0)}{r}:{L(c_a0 + W - 1)}{r})")
        ws.cell(r, c_cb, '=IF(' + "+".join(f"({c}{r}>0)" for c in pivB)
                + f'={len(PIVOTALS)},"full","CUT")')
        ws.cell(r, c_ca, '=IF(' + "+".join(f"({c}{r}>0)" for c in pivA)
                + f'={len(PIVOTALS)},"full","CUT")')
        r += 1
    want_b, want_a = sum(recon["before"].values()), sum(recon["end"].values())
    assert abs(disp_b - want_b) < 1e-6, \
        f"SOH sheet BEFORE {disp_b:,.1f} != reconciled {want_b:,.1f}"
    assert abs(disp_a - want_a) < 1e-6, \
        f"SOH sheet AFTER {disp_a:,.1f} != reconciled end {want_a:,.1f}"
    from openpyxl.utils import get_column_letter as L2
    ws.auto_filter.ref = f"A10:{L2(NC)}{r - 1}"
    widths(ws, {"A": 7, "B": 26, "C": 13, "D": 8, "E": 13, "F": 11, "G": 7, "H": 6,
                "I": 15, "J": 6, "K": 8, "L": 8, "M": 17, "N": 7})
    for k in range(W * 2):
        ws.column_dimensions[L2(c_b0 + k)].width = 5.2
    for c, w in ((c_ub, 8), (c_ua, 8), (c_cb, 8), (c_ca, 8)):
        ws.column_dimensions[L2(c)].width = w
    return name


def sheet_cutwork(wb, R, pol, m, idx, soh_name):
    H = R["holdings"]
    fl = R["fill"]
    PIVOTALS = list(pol.pivotals)
    ws = nogrid(wb.create_sheet(f"{idx} · Cut% Working"))
    NC = 7
    q = f"'{soh_name}'"
    title_block(ws, "Cut-size % — the working, step by step",
                f"a holding is CUT if it is missing ANY pivotal size · pivotals "
                f"{' · '.join(PIVOTALS)}", NC)
    r = 4
    callout(ws, r, NC,
            f"RULE: a store×style holding is CUT when any one of the {len(PIVOTALS)} "
            "pivotal sizes is absent. Miss 1–2 = Repairable · miss 3+ = Broken · "
            "miss 0 = Complete (a FULL set — frozen). Every figure below is a LIVE "
            f"formula reading the '{soh_name}' sheet.", "info")
    r += 3
    ws.cell(r, 1, "STEP 1 · the base — units and cut units, before and after").font = \
        F(11, True, NAVY)
    r += 1
    header_row(ws, r, ["Measure", "BEFORE", "AFTER", "Change", "How it is derived"])
    r += 1
    for i, (lab, fb, fa, how, fmt) in enumerate([
            ("Total units", f"={q}!C5", f"={q}!C6", "SUM of the units column", INT),
            ("Cut units", f"={q}!C7", f"={q}!C8", "SUMIF cut flag = CUT", INT),
            ("CUT-SIZE %", f"={q}!C7/{q}!C5", f"={q}!C8/{q}!C6",
             "cut units ÷ total units", PCT)]):
        datarow(ws, r, [lab, None, None, None, how], band=(i % 2 == 0),
                aligns=[LEFT, RGT, RGT, RGT, LEFT], bold=(lab == "CUT-SIZE %"))
        ws.cell(r, 2, fb).number_format = fmt
        ws.cell(r, 3, fa).number_format = fmt
        ws.cell(r, 4, f"=C{r}-B{r}").number_format = fmt
        for c in (2, 3, 4):
            ws.cell(r, c).font = F(10, True, MAROON if lab == "CUT-SIZE %" else INK)
            ws.cell(r, c).alignment = RGT
            ws.cell(r, c).border = border()
        r += 1
    r += 1
    ws.cell(r, 1, "STEP 2 · holding state — how the network splits").font = F(11, True, NAVY)
    r += 1
    header_row(ws, r, ["Holding state", "Holdings", "Units", "% of units", "meaning"])
    r += 1
    tot_u = sum(h.units for h in H.values()) or 1e-9
    for i, (lab, mean) in enumerate([
            ("Complete", "FULL sets — frozen; only whole dead sets relocate"),
            ("Repairable", "missing 1–2 pivotals — healable"),
            ("Broken", "missing 3+ — fragment, feeds the pool")]):
        n = sum(1 for h in H.values() if h.klass == lab)
        u = sum(h.units for h in H.values() if h.klass == lab)
        datarow(ws, r, [lab, n, u, u / tot_u, mean], band=(i % 2 == 0),
                aligns=[LEFT, RGT, RGT, RGT, LEFT], fmts=[None, INT, INT, PCT, None])
        r += 1
    datarow(ws, r, ["TOTAL", len(H), tot_u, 1.0,
                    "every holding in the working universe"],
            band=True, bold=True, aligns=[LEFT, RGT, RGT, RGT, LEFT],
            fmts=[None, INT, INT, PCT, None])
    r += 2
    ws.cell(r, 1, "STEP 3 · what the plan does to it").font = F(11, True, NAVY)
    r += 1
    header_row(ws, r, ["Stage", "Cut %", "Units still cut", "vs before"])
    r += 1

    def _p(s):
        return s["cut"]["ALL"] / s["tot"]["ALL"] if s["tot"]["ALL"] else 0

    base = _p(fl["before"])
    for i, (lab, st) in enumerate([
            ("BEFORE — today's SOH", fl["before"]),
            (f"after P1 — the {pol.movement_budget:,}-unit plan", fl["p1"]),
            ("AFTER — the full plan", fl["full"])]):
        datarow(ws, r, [lab, _p(st), st["cut"]["ALL"], _p(st) - base],
                band=(i % 2 == 0), aligns=[LEFT, RGT, RGT, RGT],
                fmts=[None, PCT, INT, PCT], bold=(lab.startswith("AFTER")))
        r += 1
    r += 1
    callout(ws, r, NC,
            "The AFTER column is not a projection — it is the opening SOH with every "
            "movement line applied, position by position (the end-to-end check "
            "reproduces the engine figure to six decimals). "
            f"Cut% falls {base:.1%} → {_p(fl['full']):.1%} on the full plan; the "
            f"{m['gap_struct']:,} structural gaps that remain cannot be closed by "
            "movement — the size does not exist ANYWHERE in the network. They need a "
            "fresh buy (the FS sheet shows what the warehouse could cover).", "good")
    r += 4
    widths(ws, {"A": 40, "B": 16, "C": 18, "D": 14, "E": 48, "F": 12, "G": 12})
    return ws


def sheet_recon(wb, R, pol, recon, idx, lb):
    stores, tiers_, norms = R["stores"], R["tiers"], R["norms"]
    cats = list(R["categories"])
    before, out_u, in_u = recon["before"], recon["out"], recon["inn"]
    bcat, ocat, icat = Counter(), Counter(), Counter()
    for (sh, _st), h in R["holdings"].items():
        bcat[(sh, h.cat)] += h.units
    for mv in R["moves"]:
        ocat[(mv["src"], mv["cat"])] += mv["qty"]
        icat[(mv["dst"], mv["cat"])] += mv["qty"]
    naw = R["norms_aw26"]
    fofo, fofo_net = R["fofo"], R["fofo_net"]
    ws = nogrid(wb.create_sheet(f"{idx} · Store Reconciliation"))
    NC = 24 + 3 * len(cats)
    NLB = lb["norm"]
    NCX = lb["norm_context"]
    title_block(ws, "Store reconciliation — SOH − transferred out + received = end stock",
                f"combined fill on {NLB} norms (operative) with {NCX} norms adjacent "
                "(context) · movement corridor · per-category split · franchise net flow",
                NC)
    r = 4
    tb, to, ti = sum(before.values()), sum(out_u.values()), sum(in_u.values())
    kpi(ws, r, 1, "SOH before", f"{tb:,.0f}", "units across all doors", NAVY)
    kpi(ws, r, 3, "Transferred out", f"{to:,.0f}", "from donor doors", MAROON)
    kpi(ws, r, 5, "Received", f"{ti:,.0f}", "delivered to doors", GREEN)
    kpi(ws, r, 7, "End stock in stores", f"{tb - to + ti:,.0f}",
        "equals SOH before — direct moves keep every unit in a store", TEAL)
    r += 4
    callout(ws, r, NC,
            "Every row: END = BEFORE − OUT + IN, and the Check column recomputes that "
            "in the cell. THE MOVEMENT CORRIDOR: 'Low during move' is the door's worst "
            "position — everything out, nothing landed yet — and never dips below "
            f"{pol.transit_floor:.0%} of norm ({pol.fofo_transit_floor:.0%} for a "
            "franchise door); 'Peak during move' is the other extreme — everything "
            "landed, nothing yet gone — and no CHOSEN inbound tops "
            f"{pol.recv_peak_cap:.0%} (heals exempt). END fill keeps the "
            f"{pol.fill_floor:.0%} floor per category AND combined. FRANCHISE NET "
            "FLOW: a net-ruled door's 'Net in−out' is never negative. "
            f"FILL [{NCX}] renders the same fills against the context norms — CONTEXT "
            f"ONLY, every floor and cap binds on the {NLB} norms.", "info")
    r += 4
    hdr = ([lb["store_id"], "Store", "City", "Region", "State", "Type", "L2L", "Format",
            "Tier", "Model", "SOH before", "Out", "In", "Net in−out", "END stock",
            "Check", f"{NLB} norm", f"Fill before [{NLB}]", "Low during move",
            "Peak during move", f"Fill after [{NLB}]", f"{NCX} norm",
            f"Fill before [{NCX}]", f"Fill after [{NCX}]"]
           + [h for cat in cats
              for h in (f"{cat} norm", f"{cat} fill b", f"{cat} fill a")])
    header_row(ws, r, hdr)
    hr = r
    r += 1
    universe = sorted(set(before) | set(in_u) | set(out_u),
                      key=lambda x: (-(before[x] - out_u[x] + in_u[x]),
                                     id_sort_key(x)))
    assert abs(sum(before[s] - out_u[s] + in_u[s] for s in universe)
               - (tb - to + ti)) < 1e-6
    AL = ([CTR, LEFT, LEFT, CTR, LEFT, CTR, CTR, LEFT, CTR, CTR] + [RGT] * 5 + [CTR]
          + [RGT] * (8 + 3 * len(cats)))
    FM = ([None] * 10 + [INT, INT, INT, INT, INT, None, INT, PCT, PCT, PCT, PCT,
                         INT, PCT, PCT] + [INT, PCT, PCT] * len(cats))
    for i, sh in enumerate(universe):
        st = stores.get(sh)
        ncat = {c: norms.get((sh, c), 0) for c in cats}
        n = sum(ncat.values())
        na = sum(naw.get((sh, c), 0) for c in cats)
        e = before[sh] - out_u[sh] + in_u[sh]
        catcols = []
        for c in cats:
            nc = ncat[c]
            ec = bcat[(sh, c)] - ocat[(sh, c)] + icat[(sh, c)]
            catcols += [nc or None, (bcat[(sh, c)] / nc if nc else None),
                        (ec / nc if nc else None)]
        datarow(ws, r, [idc(sh), st.name if st else "", st.city if st else "",
                        st.region if st else "", st.state if st else "",
                        st.stype if st else "", st.l2l if st else "",
                        st.fmt if st else "", tiers_.get(sh, ""),
                        st.model if st else "",
                        before[sh], out_u[sh], in_u[sh], None, e, None, n or None,
                        (before[sh] / n if n else None),
                        ((before[sh] - out_u[sh]) / n if n else None),
                        ((before[sh] + in_u[sh]) / n if n else None),
                        (e / n if n else None), na or None,
                        (before[sh] / na if na else None),
                        (e / na if na else None)] + catcols,
                band=(i % 2 == 0), aligns=AL, fmts=FM)
        cnet = ws.cell(r, 14, f"=M{r}-L{r}")
        cnet.number_format = INT
        cnet.alignment = RGT
        cnet.border = border()
        if sh in fofo:
            ws.cell(r, 10).font = F(9, True, NAVY if sh in fofo_net else AMBER)
            cnet.font = F(9, True, NAVY if sh in fofo_net else AMBER)
        c = ws.cell(r, 16, f'=IF(ABS((K{r}-L{r}+M{r})-O{r})<0.001,"OK","MISMATCH")')
        c.font = F(9, True, GREEN)
        c.alignment = CTR
        c.border = border()
        r += 1
    datarow(ws, r, ["", "NETWORK", "", "", "", "", "", "", "", "", tb, to, ti, None,
                    tb - to + ti] + [None] * (9 + 3 * len(cats)),
            band=True, bold=True, aligns=AL, fmts=FM)
    ws.cell(r, 14, f"=M{r}-L{r}").number_format = INT
    ws.cell(r, 16, f'=IF(ABS((K{r}-L{r}+M{r})-O{r})<0.001,"OK","MISMATCH")').font = \
        F(9, True, GREEN)
    from openpyxl.utils import get_column_letter as L
    ws.auto_filter.ref = f"A{hr}:{L(NC)}{r - 1}"
    widths(ws, {"A": 8, "B": 28, "C": 13, "D": 8, "E": 13, "F": 11, "G": 6, "H": 15,
                "I": 6, "J": 7, "K": 10, "L": 8, "M": 8, "N": 10, "O": 10, "P": 10,
                "Q": 9, "R": 11, "S": 11, "T": 11, "U": 11, "V": 9, "W": 11, "X": 11})
    return ws


def sheet_ros(wb, R, pol, rows, idx, lb):
    prog_label = _prog_label(pol)
    cur = lb["currency"]
    ws = nogrid(wb.create_sheet(f"{idx} · RoS Justification"))
    NC = 10
    title_block(ws, "Why these movements — the RoS gradient, move by move",
                lb["ros_note"] or "RoS = that style's recent rate of sale at that door",
                NC)
    tot = sum(x["qty"] for x in rows)
    chosen = [x for x in rows if x["program"] not in ("HEAL", "HEAL3")]
    ch_u = sum(x["qty"] for x in chosen)
    up_u = sum(x["qty"] for x in rows if x["dst_ros"] > x["src_ros"])
    w_src = sum(x["src_ros"] * x["qty"] for x in rows) / tot if tot else 0
    w_dst = sum(x["dst_ros"] * x["qty"] for x in rows) / tot if tot else 0
    r = 4
    kpi(ws, r, 1, "Units moving RoS-uphill", f"{up_u / tot:.0%}" if tot else "—",
        f"{up_u:,.0f} of {tot:,.0f} u", GREEN)
    kpi(ws, r, 3, "Mean RoS at ORIGIN", f"{w_src:.3f}", "unit-weighted, style at door", MAROON)
    kpi(ws, r, 5, "Mean RoS at DESTINATION", f"{w_dst:.3f}",
        f"{(w_dst / w_src if w_src else 0):,.1f}× the origin rate", NAVY)
    kpi(ws, r, 7, "Chosen destinations uphill", "100%",
        f"{ch_u:,.0f} u — asserted by the battery", TEAL)
    r += 4
    callout(ws, r, NC,
            "READ THIS SHEET AS THE PLAN'S DEFENCE. Every unit the engine CHOOSES a "
            "home for (SET, TOPUP, EXT) lands at a door where that style genuinely "
            f"sells — RoS ≥ {pol.min_recv_ros:.2f} — AND at least "
            f"{pol.ros_uplift:.1f}× the donor's rate. HEAL lines answer a softer bar: "
            "the destination owns the cut set, so the payoff is the completed set "
            "itself — but a heal may never run RoS-DOWNHILL.", "info")
    r += 4
    ws.cell(r, 1, "CONFIDENCE — EVERY TRANSFER GRADED").font = F(11, True, NAVY)
    r += 1
    header_row(ws, r, ["Confidence", "Meaning", "Lines", "Units",
                       f"Value {cur}".strip()])
    r += 1
    CONF_MEAN = {
        "HIGH": f"receiver RoS ≥ {pol.min_recv_ros:.2f} and ≥ "
                f"{pol.ros_uplift:.1f}× the donor's",
        "OK": "RoS-uphill but below the HIGH bar",
        "NO-DATA": "no velocity data at either end (flagged, not blocked)"}
    for i, cf in enumerate(["HIGH", "OK", "NO-DATA"]):
        px = [x for x in rows if x.get("conf") == cf]
        datarow(ws, r, [cf, CONF_MEAN[cf], len(px), sum(x["qty"] for x in px),
                        sum(x["value"] for x in px)],
                band=(i % 2 == 0), aligns=[CTR, LEFT, RGT, RGT, RGT],
                fmts=[None, None, INT, INT, INT])
        r += 1
    r += 1
    ws.cell(r, 1, "THE GRADIENT BY PROGRAM").font = F(11, True, NAVY)
    r += 1
    header_row(ws, r, ["Program", "What it does", "Lines", "Units",
                       f"Value {cur}".strip(), "RoS origin (wtd)",
                       "RoS destination (wtd)", "Uplift ×", "Uphill u", "Uphill %"])
    r += 1
    for i, p in enumerate(PROGRAMS):
        px = [x for x in rows if x["program"] == p]
        if not px:
            continue
        u = sum(x["qty"] for x in px)
        so = sum(x["src_ros"] * x["qty"] for x in px) / u
        sd = sum(x["dst_ros"] * x["qty"] for x in px) / u
        uu = sum(x["qty"] for x in px if x["dst_ros"] > x["src_ros"])
        datarow(ws, r, [p, prog_label[p], len(px), u, sum(x["value"] for x in px),
                        so, sd, (sd / so if so else None), uu, uu / u],
                band=(i % 2 == 0),
                aligns=[CTR, LEFT, RGT, RGT, RGT, RGT, RGT, RGT, RGT, RGT],
                fmts=[None, None, INT, INT, INT, "0.000", "0.000", "0.0", INT, PCT])
        r += 1
    r += 1
    ws.cell(r, 1, "DISTRIBUTION — HOW STEEP IS THE CLIMB (chosen destinations only)"
            ).font = F(11, True, NAVY)
    r += 1
    header_row(ws, r, ["RoS uplift at destination (Δ vs donor)", "Lines", "Units",
                       f"Value {cur}".strip()])
    r += 1
    for i, (lo, hi, lab) in enumerate([
            (0.0, 0.05, "+0.00 – 0.05 · marginal"), (0.05, 0.15, "+0.05 – 0.15 · solid"),
            (0.15, 0.30, "+0.15 – 0.30 · strong"),
            (0.30, 99.0, "+0.30 and above · step-change")]):
        px = [x for x in chosen if lo <= x["dst_ros"] - x["src_ros"] < hi]
        datarow(ws, r, [lab, len(px), sum(x["qty"] for x in px),
                        sum(x["value"] for x in px)],
                band=(i % 2 == 0), aligns=[LEFT, RGT, RGT, RGT],
                fmts=[None, INT, INT, INT])
        r += 1
    r += 1
    ws.cell(r, 1, "TOP 50 MOVES BY VALUE — the gradient line by line").font = \
        F(11, True, NAVY)
    r += 1
    header_row(ws, r, ["Priority", "Program", "Style", "From", "From door",
                       "From profile", "To", "To door", "To profile", "Units",
                       "RoS from → to", f"Value {cur}".strip()])
    r += 1
    stores = R["stores"]

    def _prof(sh):
        s = stores.get(sh)
        return (f"{s.region} · {s.state} · {s.stype} · {s.l2l} · {s.fmt}" if s else "")

    grouped = defaultdict(lambda: dict(qty=0.0, value=0.0))
    for x in chosen:
        g = grouped[(x["band"], x["program"], x["style"], x["src"], x["dst"],
                     x["src_ros"], x["dst_ros"])]
        g["qty"] += x["qty"]
        g["value"] += x["value"]
    top = sorted(grouped.items(), key=lambda kv: -kv[1]["value"])[:50]
    for i, ((band, prog, style, src, dst, sr, dr), g) in enumerate(top):
        datarow(ws, r, [band, prog, style, idc(src), stores[src].name, _prof(src),
                        idc(dst), stores[dst].name, _prof(dst),
                        g["qty"], f"{sr:.2f} → {dr:.2f}", g["value"]],
                band=(i % 2 == 0),
                aligns=[CTR, CTR, LEFT, CTR, LEFT, LEFT, CTR, LEFT, LEFT, RGT, CTR, RGT],
                fmts=[None, None, None, None, None, None, None, None, None, INT,
                      None, INT])
        ws.cell(r, 1).font = F(9, True, BAND_COL[band])
        r += 1
    r += 1
    callout(ws, r, NC,
            "Sheet 1 carries 'RoS from' and 'RoS to' on EVERY line — filter any style, "
            "door or program there to see its slice of this gradient. Where a style "
            "has no recorded rate at either end the donor is dead stock (RoS 0) moving "
            "to a live seller, the cleanest uplift there is.", "good")
    r += 4
    widths(ws, {"A": 12, "B": 12, "C": 18, "D": 8, "E": 26, "F": 8, "G": 26, "H": 8,
                "I": 14, "J": 12})
    return ws


def sheet_fs(wb, R, pol, idx, lb):
    fs, rep = R["fs"], R["fs_report"]
    cur = lb["currency"]
    whs = list(pol.fs_warehouses) or sorted({w for d in fs.values() for w in d})
    ws = nogrid(wb.create_sheet(f"{idx} · FS — Warehouse Stock"))
    NC = 6 + len(whs)
    title_block(ws, "FS — warehouse free stock. REPORT ONLY, not part of the movement plan",
                "the movement budget is store-to-store; releasing FS is a separate, "
                "explicit decision"
                + (f" · {lb['fs_note']}" if lb["fs_note"] else ""), NC)
    need_cov = sum(min(x["need"], x["fs_total"]) for x in rep)
    r = 4
    for i, w in enumerate(whs[:3]):
        tot_w = sum(d.get(w, 0) for d in fs.values())
        kpi(ws, r, 1 + i * 2, f"FS on hand — {w}", f"{tot_w:,.0f} u", "free stock", NAVY)
    kpi(ws, r, 1 + min(len(whs), 3) * 2, "Movement gaps FS could close",
        f"{len(rep):,}", "style×size lines", TEAL)
    kpi(ws, r, 3 + min(len(whs), 3) * 2, "Units of gap FS covers",
        f"{need_cov:,.0f}", "if released — the buyer's call", GREEN)
    r += 4
    callout(ws, r, NC,
            "ZERO lines of the movement plan draw on this stock. The plan's remaining "
            "gaps are sizes movement cannot supply — no store donor exists. The table "
            "matches those gaps against what sits in the warehouses TODAY: releasing "
            "any of it is a separate, explicit decision that would close the listed "
            "gap without consuming a unit of the store-to-store budget.", "warn")
    r += 4
    ws.cell(r, 1, "GAPS THE WAREHOUSE COULD CLOSE").font = F(11, True, NAVY)
    r += 1
    header_row(ws, r, ["Style", "Cat", "Size", "Gap u (doors waiting)", "FS total"]
               + [f"at {w}" for w in whs] + [f"MRP {cur}".strip()])
    r += 1
    hr = r
    for i, x in enumerate(rep):
        datarow(ws, r, [x["style"], x["cat"], x["size"], x["need"], x["fs_total"]]
                + [x["wh"].get(w, 0.0) for w in whs] + [x["mrp"]],
                band=(i % 2 == 0),
                aligns=[LEFT, CTR, CTR, RGT, RGT] + [RGT] * len(whs) + [RGT],
                fmts=[None, None, None, INT, INT] + [INT] * len(whs) + [INT])
        r += 1
    if not rep:
        datarow(ws, r, ["— no open gap is serviceable from FS —"]
                + [None] * (NC - 1))
        r += 1
    from openpyxl.utils import get_column_letter as L
    ws.auto_filter.ref = f"A{hr - 1}:{L(5 + len(whs) + 1)}{r - 1}"
    widths(ws, {"A": 18, "B": 9, "C": 7, "D": 18, "E": 10, "F": 9, "G": 9, "H": 10})
    return ws


def sheet_checks(wb, battery, rows):
    ws = nogrid(wb.create_sheet("8 · Checks"))
    NC = 6
    checks = battery.checks
    title_block(ws, f"Pre-execution checks — all {len(checks)}, from every angle",
                "this file instructs physical movement that cannot be undone", NC)
    r = 4
    p, f_, s = battery.counts()
    kpi(ws, r, 1, "Checks", battery.summary(),
        "build refuses to save on any failure", GREEN if not f_ else MAROON)
    kpi(ws, r, 3, "Lines in the file", f"{len(rows):,}",
        f"{sum(x['qty'] for x in rows):,.0f} units", NAVY)
    kpi(ws, r, 5, "Replayed from the file", "yes",
        "checks read the rows, not the engine", TEAL)
    r += 4
    sections = []
    for c in checks:
        if c.section not in sections:
            sections.append(c.section)
    n = 0
    for sec in sections:
        ws.cell(r, 1, sec.upper()).font = F(11, True, NAVY)
        r += 1
        header_row(ws, r, ["#", "Id", "Check", "Result", "Evidence"])
        r += 1
        for i, c in enumerate([c for c in checks if c.section == sec]):
            n += 1
            state = ("SKIPPED" if c.skipped
                     else "PASS" if c.passed
                     else "AMBER" if c.warn else "FAIL")
            col = (MUTE if c.skipped else GREEN if c.passed
                   else AMBER if c.warn else MAROON)
            datarow(ws, r, [n, c.id, c.name, state, c.detail],
                    band=(i % 2 == 0), aligns=[CTR, LEFT, LEFT, CTR, LEFT],
                    fmts=[INT, None, None, None, None], fontc=col, bold=True)
            r += 1
        r += 1
    callout(ws, r, NC,
            "EVERY check blocks the build: if one fails no workbook is written, so a "
            "file that exists is a file that passed. Checks are computed from the ROWS "
            "OF THIS FILE replayed onto the opening SOH — not from engine state — so "
            "the file cannot drift from what was verified. A SKIPPED row is a rule "
            "disabled by policy: its check is shown, not silently dropped, so a "
            "shrunk battery can never pass for a passing one.", "good")
    r += 4
    widths(ws, {"A": 5, "B": 18, "C": 52, "D": 9, "E": 84, "F": 12})
    return ws


def sheet_citydist(wb, R, pol, idx):
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter as L
    stores = R["stores"]
    aliases = dict(pol.city_aliases)
    pts, nogeo = {}, set()
    for sh in sorted(R["sb_stores"], key=id_sort_key):
        s = stores[sh]
        c = canon_city(s.city, aliases)
        if not c:
            continue
        if s.has_geo:
            pts.setdefault(c, []).append((s.lat, s.lon))
        else:
            nogeo.add(c)
    cent = {c: (sum(p[0] for p in ps) / len(ps), sum(p[1] for p in ps) / len(ps))
            for c, ps in pts.items()}
    cities = sorted(cent)
    missing = sorted(nogeo - set(cent))
    ws = nogrid(wb.create_sheet(f"{idx} · City Distances (km)"))
    title_block(ws, "CITY-TO-CITY DISTANCES — straight-line km between every network city",
                f"{len(cities)} cities holding stock · city point = centroid of its "
                "stores · these are the distances the donor localization minimises", 8)
    callout(ws, 4, 8,
            "Read row city → column city. Straight-line (haversine) km — road distance "
            "runs longer. Reference only: the movement sheets are already paired to "
            "the nearest legal donor per style×size."
            + (f" No coordinates in the master for: {', '.join(missing)}."
               if missing else ""), "info")
    hr = 8
    rot = Alignment(textRotation=90, horizontal="center", vertical="bottom")
    ws.cell(hr, 1, "City").font = F(9, True, SLATE)
    for j, c in enumerate(cities):
        cell = ws.cell(hr, 2 + j, c)
        cell.font = F(8, True, SLATE)
        cell.alignment = rot
        ws.column_dimensions[L(2 + j)].width = 5.4
    ws.row_dimensions[hr].height = 95
    for i, ca in enumerate(cities):
        ws.cell(hr + 1 + i, 1, ca).font = F(8, True, INK)
        (la, lo) = cent[ca]
        for j, cb in enumerate(cities):
            if j < i:
                continue
            km = 0.0 if i == j else haversine(la, lo, *cent[cb])
            for rr, cc in (((hr + 1 + i), (2 + j)), ((hr + 1 + j), (2 + i))):
                cell = ws.cell(rr, cc, round(km))
                cell.number_format = "#,##0"
                cell.font = F(8, False, MUTE if km else SLATE)
    ws.column_dimensions["A"].width = 24
    return ws


def sheet_readme(wb, R, pol, rows, m, battery, lb):
    samecity_rows = [x for x in rows
                     if x.get('same_city') and x['program'] != 'HEAL3']
    ws = nogrid(wb.create_sheet("0 · Read Me", 0))
    NC = 6
    w2lbl = "·".join(pol.wave2_pivotals)
    subtitle_bits = [b for b in (f"SOH {lb['soh_asof']}" if lb["soh_asof"] else "",
                                 lb["ros_note"]) if b]
    title_block(ws, lb["title"], " · ".join(subtitle_bits) or "direct store-to-store",
                NC)
    r = 4
    for t in [
        "EVERY LINE IS ONE PHYSICAL TRANSFER: take the unit from the door in 'Take "
        "from', deliver it to the door in 'Deliver to'. No warehouse legs, no hub "
        "holds, no routing points — the donor named on the line hands the unit over.",
        "FULL SETS ARE FROZEN. No article that is part of a full size set moves "
        "individually. The one whole-set exception: a DEAD full set relocates INTACT "
        "to a door that sells the style.",
        "EVERY CHOSEN DESTINATION CLIMBS THE RoS GRADIENT. A unit lands in a cut set "
        "to complete it, or in a full-set door whose style RoS strictly beats the "
        "donor's.",
        "WAREHOUSE FREE STOCK IS NOT IN THIS PLAN — it appears only on the FS sheet "
        "as a separate report. The movement budget is store-to-store by construction.",
    ]:
        datarow(ws, r, [t] + [None] * (NC - 1), aligns=[LEFT] * NC)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
        r += 1
    r += 1
    ws.cell(r, 1, "WHAT THIS PLAN DOES").font = F(11, True, NAVY)
    r += 1
    header_row(ws, r, ["Metric", "Value"])
    r += 1
    a, nn = R["afs_stat"], R["nonorm_stat"]
    ls = R["loc_stat"]
    w2, am = R["wave2_stat"], R["asm_stat"]
    loc_val = (f"{ls['lines_repaired']:,} lines re-paired · "
               f"{ls['km_before'] - ls['km_after']:,.0f} km saved"
               if ls["km_before"] else "—")
    gap3 = (f"{w2['cut3_open'][0] / w2['cut3_open'][1]:.1%} → "
            f"{w2['cut3_w1'][0] / w2['cut3_w1'][1]:.1%} → "
            f"{w2['cut3_end'][0] / w2['cut3_end'][1]:.1%}"
            if w2["cut3_open"][1] else "—")
    for i, (lab, v, f2) in enumerate([
            ("Cut units today", m["cut_u"], INT),
            ("Cut % today", m["cut_pct_u"], PCT),
            (f"Cut % after the P1 {pol.movement_budget:,}-unit plan", m["cut_p1"], PCT),
            ("Cut % after the full plan", m["cut_full"], PCT),
            ("Cut sets resolved — full plan", m["resolved"], INT),
            ("Frozen-style units untouched", m["core_u"], INT),
            ("Frozen doors — no donate, no receive",
             f"{a['doors']} doors · {a['units']:,.0f} u", None),
            ("No-norm doors evacuated to good stores",
             f"{nn['doors']} doors · {nn['out']:,.0f} u out, 0 in", None),
            ("Master-frozen doors — blocked / not-open / named model",
             f"{R['frozen_stat']['doors']} doors · "
             f"{R['frozen_stat']['units']:,.0f} u untouched", None),
            ("New doors — donor-frozen",
             f"{R['newdoor_stat']['doors']} doors · "
             f"{R['newdoor_stat']['units']:,.0f} u held · received "
             f"{R['newdoor_stat']['inn']:,.0f} u", None),
            ("No-norm franchise doors — EXEMPT, stock untouched",
             f"{R['nnf_stat']['doors']} doors · {R['nnf_stat']['units']:,.0f} u "
             "frozen in place", None),
            ("Franchise net flow — receive ≥ give",
             f"{R['fofo_stat']['net_doors']} doors · out "
             f"{R['fofo_stat']['out']:,.0f} ≤ in {R['fofo_stat']['inn']:,.0f} u", None),
            ("Same-city transfers — split to sheet 1b",
             f"{len(samecity_rows):,} lines · "
             f"{sum(x['qty'] for x in samecity_rows):,.0f} u",
             None),
            ("Donors localized — nearest legal donor per style×size", loc_val, None),
            (f"Wave 2 — new {w2lbl} sets on the after-SOH (sheet 1c)",
             f"{w2['sets']:,} sets · {w2['units']:,.0f} u", None),
            ("Assembly — NEW sets at dept-selling doors (sheet 1d)",
             f"{am['sets']:,} sets · {am['units']:,.0f} u · {am['doors']:,} doors",
             None),
            (f"{w2lbl} availability gap open → wave 1 → wave 2", gap3, None),
            ("Units to move in P1", m["p1_units"], INT),
            ("Lines to execute in P1", m["p1_lines"], INT),
            ("Units moving to a strictly better-RoS door", m["uphill_u"], INT),
            ("Units at HIGH RoS-confidence", m["conf_hi"], INT),
            ("Structural gaps — need a fresh buy (national)", m["gap_struct"], INT),
            ("Full sets relocated whole", m["reloc_n"], INT),
            ("Heals GEO-BLOCKED by the transfer geography",
             R.get("geo_stat", {}).get("blocked_sets", 0), INT),
            ("Checks", battery.summary(), None)]):
        datarow(ws, r, [lab, v], band=(i % 2 == 0), aligns=[LEFT, RGT],
                fmts=[None, f2])
        r += 1
    r += 1
    callout(ws, r, NC,
            "THE SHEETS — 1 · Movement: every wave-1 INTER-CITY transfer, filter "
            f"Priority = P1 to execute the {pol.movement_budget:,}-unit plan. "
            "1b · Same-City Moves: wave-1 transfers where both doors share a city. "
            f"1c · Wave 2: {w2lbl} set completion on the plan's AFTER position — "
            "execute only once 1/1b have landed. 1d · Assembly: NEW sets built at "
            "dept-selling doors from dead stock. 2 · Fill Rate: cut% and fill before "
            "/ P1 / after. 3 · SOH: every store × style × size before and after. "
            "4 · Cut% Working: the derivation, live formulas. 5 · Store "
            "Reconciliation: SOH − out + in = end per door, with the movement "
            "corridor and franchise net flow. 6 · RoS Justification: the plan seen "
            "purely through RoS. 7 · FS — warehouse stock, REPORT ONLY. 8 · Checks: "
            "the full battery with evidence. 9 · City Distances: straight-line km "
            "between every pair of network cities.", "info")
    r += 4
    ws.cell(r, 1, "THE RULES IN FORCE (from the rule registry)").font = F(11, True, NAVY)
    r += 1
    header_row(ws, r, ["Rule", "Status", "What it enforces"])
    r += 1
    for i, rl in enumerate(RULES):
        on = bool(rl.enabled(pol))
        datarow(ws, r, [rl.title, "ACTIVE" if on else "OFF", rl.doc],
                band=(i % 2 == 0), aligns=[LEFT, CTR, LEFT],
                fontc=(INK if on else MUTE))
        ws.cell(r, 2).font = F(9, True, GREEN if on else MUTE)
        r += 1
    r += 1
    widths(ws, {"A": 62, "B": 20, "C": 90, "D": 16, "E": 16, "F": 16})
    return ws


# ───────────────────────── entry point ─────────────────────────
def render_workbook(R, out_path, labels=None):
    """Battery first, then the 13 sheets, then preflighted save. Returns the
    battery. Raises before writing anything if any check fails — a workbook
    that exists is a workbook that passed."""
    pol = R["policy"]
    lb = dict(DEFAULT_LABELS, **(labels or {}))
    rows = enrich_rows(R, pol, lb)
    recon = build_recon(R)
    battery = run_battery(R, rows, recon, pol)
    if battery.failed():
        lines = "\n".join(f"  {c.id}: {c.name} — {c.detail}" for c in battery.failed())
        raise AssertionError(f"build blocked — failed checks:\n{lines}")
    m = _metrics(R, rows)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheet_moves(wb, R, pol, rows, lb)
    sheet_samecity(wb, R, pol, rows, lb)
    sheet_wave2(wb, R, pol, rows, lb)
    sheet_assembly(wb, R, pol, rows, lb)
    sheet_fill(wb, R, pol, m, lb)
    soh_name = sheet_soh(wb, R, pol, recon, 3, lb)
    sheet_cutwork(wb, R, pol, m, 4, soh_name)
    sheet_recon(wb, R, pol, recon, 5, lb)
    sheet_ros(wb, R, pol, rows, 6, lb)
    sheet_fs(wb, R, pol, 7, lb)
    sheet_checks(wb, battery, rows)
    sheet_citydist(wb, R, pol, 9)
    sheet_readme(wb, R, pol, rows, m, battery, lb)
    safe_save(wb, out_path)
    return battery
