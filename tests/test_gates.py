"""Gates the review proved missing: the public plan-content golden, the red
blocking path, sanity finding units, loader-semantics units, the degraded
configuration end to end, and warn semantics.
"""
import atexit
import json
import shutil
import tempfile
from pathlib import Path

import openpyxl
import yaml

from restitch.core import engine
from restitch.core.checks import Battery
from restitch.core.model import EngineInputs, Holding, Store
from restitch.core.policy import from_dict
from restitch.core.summary import canonical_moves
from restitch.io.manifest import load_manifest, resolve
from restitch.io.sanity import run_sanity
from restitch.synth import write_demo

FIXTURES = Path(__file__).resolve().parent / "fixtures"

_CACHE: dict = {}


def _tmp(prefix):
    d = Path(tempfile.mkdtemp(prefix=prefix))
    atexit.register(shutil.rmtree, d, ignore_errors=True)
    return d


def _demo():
    if "demo" not in _CACHE:
        d = _tmp("restitch-gates-")
        man = write_demo(d, seed=8)
        _CACHE["demo"] = (d, man)
    return _CACHE["demo"]


# ── the public plan-content golden ──────────────────────────────────────
def test_synth_golden_seed8_moves_exact():
    """The tie-break mutant survived the whole public suite because no public
    test pinned plan CONTENT (review test-rigor M3/#5). This is the OSS tree's
    golden: same bundle, same policy, byte-same canonical moves."""
    d, man = _demo()
    rr = resolve(load_manifest(man),
                 from_dict(yaml.safe_load((d / "policy.yaml").read_text())))
    got = [list(t) for t in canonical_moves(engine.run(rr.inputs, rr.policy))]
    g = json.loads((FIXTURES / "synth_golden_seed8.json").read_text())
    assert len(got) == g["lines"]
    if got != g["moves"]:
        gs = {tuple(t) for t in got}
        ws = {tuple(t) for t in g["moves"]}
        raise AssertionError(
            f"public plan drift — missing {sorted(ws - gs)[:4]} · "
            f"extra {sorted(gs - ws)[:4]}")


# ── the red blocking path (CLI) ─────────────────────────────────────────
def test_red_sanity_blocks_the_cli_run():
    """'RED blocks the run' was the product's headline behavior with no
    regression net anywhere (review test-rigor #6)."""
    from restitch.cli import main
    d, _man = _demo()
    poisoned = _tmp("restitch-red-")
    shutil.copytree(d, poisoned, dirs_exist_ok=True)
    prof = yaml.safe_load((poisoned / "mappings" / "soh.yaml").read_text())
    prof["semantics"]["category_values"] = {"Jacket": ["NOPE"], "Trouser": ["NADA"]}
    (poisoned / "mappings" / "soh.yaml").write_text(yaml.safe_dump(prof))
    rc = main(["run", "--manifest", str(poisoned / "manifest.yaml"),
               "--policy", str(poisoned / "policy.yaml"),
               "--out", str(poisoned / "run")])
    assert rc == 2, "RED findings must block with exit code 2"
    assert "no_stock" in (poisoned / "run" / "sanity.txt").read_text()
    assert not (poisoned / "run" / "plan.json").exists(), \
        "a blocked run must produce no plan"


# ── sanity finding units ────────────────────────────────────────────────
def _tiny_inputs(mrp=999.0, cities=None):
    cities = cities or {1: ("MUMBAI", "MH"), 2: ("DELHI", "DL")}
    stores = {sh: Store(shrm=sh, name=f"S{sh}", city=c, state=st)
              for sh, (c, st) in cities.items()}
    holdings = {(sh, "S"): Holding(shrm=sh, style="S", cat="J", div="M",
                                   sizes={"096": 2.0}, mrp=mrp)
                for sh in stores}
    return EngineInputs(stores=stores, tiers={}, holdings=holdings, excluded={},
                        dept_votes={}, perf={}, style_net={},
                        norms={(sh, "J"): 2.0 for sh in stores},
                        norms_context={}, fs={}, fs_meta={})


def _pol():
    from restitch.core.policy import Policy
    return Policy(all_sizes=("096", "100"), pivotals=("096", "100"),
                  wave2_pivotals=(), assembly_enabled=False)


def test_extended_money_is_a_red_finding():
    findings = run_sanity(_tiny_inputs(mrp=431766.0), _pol())
    hit = [f for f in findings if f.id == "mrp_extended"]
    assert hit and hit[0].level == "red"


def test_city_near_miss_fires_and_an_alias_silences_it():
    inp = _tiny_inputs(cities={1: ("BANGALORE", "KA"), 2: ("BENGALURU", "KA"),
                              3: ("BANGALORS", "KA")})
    hit = [f for f in run_sanity(inp, _pol()) if f.id == "city_near_miss"]
    assert hit, "one-and-two-edit spellings in one state must warn"
    aliased = _pol().with_(city_aliases=(("BENGALURU", "BANGALORE"),
                                         ("BANGALORS", "BANGALORE")))
    assert not [f for f in run_sanity(inp, aliased) if f.id == "city_near_miss"]


def test_export_row_cap_detected():
    findings = run_sanity(_tiny_inputs(), _pol(),
                          raw_counts={("soh", "x.xlsx", "SOH"): 150000})
    assert any(f.id == "export_row_cap" for f in findings)


def test_warn_checks_report_amber_and_never_block():
    b = Battery()
    b.add("w.only", "s", "W", False, "advisory", warn=True)
    assert b.ok and not b.failed() and [c.id for c in b.warned()] == ["w.only"]


# ── loader semantics units (review test-rigor #7) ───────────────────────
def _wb(path, sheet, header, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(header)
    for r in rows:
        ws.append(r)
    wb.save(path)


def test_ros_overlays_refresh_creates_and_authoritative_clamps():
    from restitch.io.loaders import load_ros
    from restitch.io.mapping import RoleMapping
    d = _tmp("restitch-ros-")
    _wb(d / "base.xlsx", "V", ["Store", "Style", "R8", "S8"],
        [[1, "A", 0.2, 3]])
    _wb(d / "ref.xlsx", "V", ["Store", "Style", "R8", "R4"],
        [[1, "A", 0.4, 0.1], [2, "B", 0.3, 0.2]])       # (2,B) is NEW
    _wb(d / "auth.xlsx", "V", ["Store", "Style", "R8", "R4", "S8", "S4", "G8"],
        [[1, "A", -0.5, -0.1, 7, 3, 9]])                # negative RoS clamps
    base = RoleMapping(role="ros", name="b", sheet="V", header_row=0,
                       columns=dict(location="Store", style="Style",
                                    ros8="R8", sales8="S8"),
                       semantics=dict(kind="base"))
    ref = RoleMapping(role="ros", name="r", sheet="V", header_row=0,
                      columns=dict(location="Store", style="Style",
                                   ros8="R8", ros4="R4"),
                      semantics=dict(kind="refresh"))
    auth = RoleMapping(role="ros", name="a", sheet="V", header_row=0,
                       columns=dict(location="Store", style="Style", ros8="R8",
                                    ros4="R4", sales8="S8", sales4="S4",
                                    grn8="G8"),
                       semantics=dict(kind="authoritative",
                                      evidence_prefix="d6_"))
    perf, _net = load_ros([(d / "base.xlsx", base), (d / "ref.xlsx", ref),
                           (d / "auth.xlsx", auth)])
    assert ("2", "B") in perf, "refresh must CREATE records the base never had"
    rec = perf[("1", "A")]      # store ids are canonical strings (core.ids)
    assert rec["ros8"] == 0.0 and rec["ros4"] == 0.0, "authoritative clamps negatives"
    assert rec["d6_sold8"] == 7 and rec["d6_grn8"] == 9
    assert rec["sales8"] == 3, \
        "evidence stays in prefixed fields — base sales (liveness) untouched"


def test_load_fs_fallback_ladder():
    from restitch.io.loaders import load_fs
    from restitch.io.mapping import RoleMapping
    d = _tmp("restitch-fs-")
    _wb(d / "fs.xlsx", "TOT",
        ["Style", "092", "096", "Grand Total"],
        [["A", 4, 0, 4], ["B", 0, 6, 6], ["C", 2, 0, 2]])
    wb = openpyxl.load_workbook(d / "fs.xlsx")
    ws = wb.create_sheet("SPLIT")
    ws.append(["Style", "Plant", "Size", "Qty", "Brand", "Category"])
    ws.append(["A", "W1", "092", 3, "X", "Cat"])        # (style,size) split
    ws.append(["A", "W2", "092", 1, "X", "Cat"])
    ws.append(["B", "W1", "092", 5, "X", "Cat"])        # style-level only
    wb.save(d / "fs.xlsx")
    tot = RoleMapping(role="fs_totals", name="t", sheet="TOT", header_row=0,
                      columns={}, semantics=dict(terminator_header="Grand Total",
                                                 style_col_index=0))
    spl = RoleMapping(role="fs_split", name="s", sheet="SPLIT", header_row=0,
                      columns=dict(style="Style", plant="Plant", size="Size",
                                   qty="Qty", brand="Brand", category="Category"))
    fs, _meta = load_fs(d / "fs.xlsx", tot, spl, all_sizes=("092", "096"),
                        fallback_plants=("W1", "W2"))
    assert abs(fs[("A", "092")]["W1"] - 3.0) < 1e-9      # exact split honored
    assert abs(fs[("B", "096")]["W1"] - 6.0) < 1e-9      # style fallback
    assert abs(fs[("C", "092")]["W1"] - 2 * 8 / 9) < 1e-9  # overall ratio (8:1)


# ── the degraded configuration, end to end ──────────────────────────────
def test_wave2_off_pipeline_surfaces_skips_everywhere():
    from restitch.core.verify import verify_workbook
    from restitch.render.workbook import render_workbook
    d, man = _demo()
    base = yaml.safe_load((d / "policy.yaml").read_text())
    rr = resolve(load_manifest(man),
                 from_dict({**base, "wave2_pivotals": [],
                            "assembly_enabled": False}))
    R = engine.run(rr.inputs, rr.policy)
    assert not any(m["program"] in ("HEAL3", "ASM") for m in R["moves"])
    out = _tmp("restitch-degraded-") / "wb.xlsx"
    battery = render_workbook(R, out)
    skipped = {c.id for c in battery.checks if c.skipped}
    assert {"wave2_replay", "assembly_replay"} <= skipped
    vb = verify_workbook(out, rr.inputs, rr.policy)
    assert vb.ok
    assert {"v_wave2", "v_assembly"} <= {c.id for c in vb.checks if c.skipped}


if __name__ == "__main__":
    import sys
    mod = sys.modules["__main__"]
    for name in sorted(n for n in dir(mod) if n.startswith("test_")):
        getattr(mod, name)()
        print(f"{name}: OK")
