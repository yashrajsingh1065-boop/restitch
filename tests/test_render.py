"""Renderer over the synthetic bundle — 13 sheets, battery-first, skip surfacing.

The private overlay proves cell-exact parity against the legacy renderer on
real data; these tests prove the public path: a synth run renders end to end,
the battery blocks before anything is written, disabled rules surface as
skipped checks in place, and every rule's declared check ids actually exist
in the battery.
"""
import tempfile
from pathlib import Path

import openpyxl
import yaml

from restitch.core import engine
from restitch.core.rules import RULES

_CACHE: dict = {}


def _rendered():
    if "out" not in _CACHE:
        from restitch.cli import main
        from restitch.core.policy import from_dict
        from restitch.io.manifest import load_manifest, resolve
        from restitch.render.workbook import render_workbook
        d = Path(tempfile.mkdtemp(prefix="restitch-render-"))
        import atexit
        import shutil
        atexit.register(shutil.rmtree, d, ignore_errors=True)
        assert main(["demo", "--out", str(d)]) == 0
        rr = resolve(load_manifest(d / "manifest.yaml"),
                     from_dict(yaml.safe_load((d / "policy.yaml").read_text())))
        R = engine.run(rr.inputs, rr.policy)
        out = d / "movement.xlsx"
        battery = render_workbook(R, out, labels=dict(soh_asof="2026-07-15"))
        _CACHE.update(out=out, battery=battery, R=R)
    return _CACHE


def test_renders_thirteen_sheets():
    c = _rendered()
    wb = openpyxl.load_workbook(c["out"], read_only=True)
    assert len(wb.sheetnames) == 13
    assert wb.sheetnames[0] == "0 · Read Me"
    assert "8 · Checks" in wb.sheetnames
    assert all(len(n) <= 31 for n in wb.sheetnames)


def test_battery_passes_and_surfaces_skips():
    b = _rendered()["battery"]
    assert b.ok
    skipped = {c.id for c in b.checks if c.skipped}
    # demo tenant lists are empty and geography is open — those rules skip IN
    # PLACE; everything else computes and passes
    assert skipped == {"core_freeze", "afs_freeze", "geo_edges"}
    p, f, s = b.counts()
    assert (f, s) == (0, 3) and p >= 35


def test_every_declared_check_id_is_in_the_battery():
    ids = {c.id for c in _rendered()["battery"].checks}
    for r in RULES:
        for cid in r.check_ids:
            assert cid in ids, f"rule {r.id} declares {cid} but the battery never emits it"


def test_failed_check_blocks_the_save():
    import pytest

    from restitch.render.workbook import render_workbook
    c = _rendered()
    R = dict(c["R"])
    R["moves"] = list(R["moves"])
    bad = dict(R["moves"][0])
    bad["qty"] = 0.5                      # non-integer qty: executable-lines check fails
    R["moves"].append(bad)
    out = Path(tempfile.mkdtemp(prefix="restitch-blocked-")) / "blocked.xlsx"
    with pytest.raises(AssertionError, match="build blocked"):
        render_workbook(R, out)
    assert not out.exists(), "a failed battery must never leave a workbook behind"


if __name__ == "__main__":
    import sys
    mod = sys.modules["__main__"]
    for name in sorted(n for n in dir(mod) if n.startswith("test_")):
        getattr(mod, name)()
        print(f"{name}: OK")
