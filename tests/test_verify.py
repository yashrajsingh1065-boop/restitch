"""Independent verifier + the footwear-UK generalization gate.

The verifier's independence is STRUCTURAL, not aspirational: an AST test walks
its imports and fails the suite if it ever touches the engine, the localizer,
the io stack or the renderer. Its checks replay the SAVED workbook against
independently loaded inputs — including a hand-corrupted artifact, which it
must catch.

footwear-UK: the same pipeline on a half-size shoe ladder written as floats —
mapping -> uk_shoe normalizer -> engine -> render -> verify, no tailoring
assumption surviving anywhere.
"""
import ast
import shutil
import tempfile
from pathlib import Path

import pytest
import yaml

from restitch.core import engine
from restitch.core.rules import RULES

_CACHE: dict = {}


def _bundle(vocab):
    key = f"b_{vocab}"
    if key not in _CACHE:
        from restitch.cli import main
        from restitch.core.policy import from_dict
        from restitch.io.manifest import load_manifest, resolve
        from restitch.render.workbook import render_workbook
        d = Path(tempfile.mkdtemp(prefix=f"restitch-{vocab}-"))
        import atexit
        import shutil
        atexit.register(shutil.rmtree, d, ignore_errors=True)
        assert main(["demo", "--out", str(d), "--vocab", vocab]) == 0
        rr = resolve(load_manifest(d / "manifest.yaml"),
                     from_dict(yaml.safe_load((d / "policy.yaml").read_text())))
        R = engine.run(rr.inputs, rr.policy)
        out = d / "wb.xlsx"
        render_workbook(R, out)
        _CACHE[key] = dict(dir=d, rr=rr, R=R, out=out)
    return _CACHE[key]


# ── structural independence ─────────────────────────────────────────────
_FORBIDDEN = ("engine", "localize", "summary")     # core modules the verifier may not touch


def test_verifier_imports_are_independent():
    vdir = Path(__file__).resolve().parent.parent / "restitch" / "core" / "verify"
    for py in vdir.glob("*.py"):
        tree = ast.parse(py.read_text())
        for node in ast.walk(tree):
            names = []
            if isinstance(node, ast.Import):
                names = [a.name for a in node.names]
            elif isinstance(node, ast.ImportFrom):
                mod = node.module or ""
                lvl = node.level
                if lvl:                    # relative: resolve against core.verify
                    base = {1: "restitch.core.verify", 2: "restitch.core",
                            3: "restitch"}[lvl]
                    mod = f"{base}.{mod}" if mod else base
                names = [mod]
            for name in names:
                if name.split(".")[-1] == "io" or ".io." in name or name.endswith(".io"):
                    raise AssertionError(f"{py.name} imports the io stack: {name}")
                if "render" in name and "restitch" in name:
                    raise AssertionError(f"{py.name} imports the renderer: {name}")
                for bad in _FORBIDDEN:
                    if name.endswith(f".{bad}"):
                        raise AssertionError(
                            f"{py.name} imports restitch.core.{bad} — the verifier "
                            "must re-derive, never reuse")


def test_every_rule_verify_id_exists_in_the_verifier():
    src = (Path(__file__).resolve().parent.parent
           / "restitch" / "core" / "verify" / "replay.py").read_text()
    for r in RULES:
        for vid in r.verify_ids:
            assert f'"{vid}"' in src, f"rule {r.id} declares {vid}, verifier never emits it"


# ── verify over the rendered synth workbook ─────────────────────────────
def test_verify_green_on_tailoring():
    from restitch.core.verify import verify_workbook
    c = _bundle("tailoring")
    vb = verify_workbook(c["out"], c["rr"].inputs, c["rr"].policy)
    assert vb.ok
    skipped = {ck.id for ck in vb.checks if ck.skipped}
    assert skipped == {"v_core", "v_afs", "v_geo"}, skipped


def _forged(c, mutate):
    import openpyxl
    bad = Path(tempfile.mkdtemp(prefix="restitch-forge-")) / "bad.xlsx"
    shutil.copy(c["out"], bad)
    wb = openpyxl.load_workbook(bad)
    ws = wb["1 · Movement — Take & Deliver"]
    hdr = data = None
    for r, row in enumerate(ws.iter_rows(values_only=True), 1):
        if hdr is None and row and row[0] == "Priority":
            hdr = ({str(h).strip(): j + 1 for j, h in enumerate(row)
                    if h is not None}, r)
        elif hdr and row[0] in ("P1", "P2", "P3") and data is None:
            data = r
    mutate(ws, hdr[0], hdr[1], data)
    wb.save(bad)
    return bad


def _append_heal(ws, ix, data_row, *, src, dst, style, size):
    r = ws.max_row + 1
    for j in range(1, len(ix) + 1):          # clone a real row's shape
        ws.cell(r, j, ws.cell(data_row, j).value)
    for name, v in (("Priority", "P1"), ("Program", "HEAL"), ("Take from", src),
                    ("Deliver to", dst), ("Style", style), ("Size", size),
                    ("Qty", 1), ("RoS from", 0.0), ("RoS to", 0.5)):
        ws.cell(r, ix[name], v)


def test_verify_catches_an_inflated_qty():
    from restitch.core.verify import verify_workbook
    c = _bundle("tailoring")
    bad = _forged(c, lambda ws, ix, hr, dr:
                  ws.cell(dr, ix["Qty"],
                          float(ws.cell(dr, ix["Qty"]).value) + 5))
    with pytest.raises(AssertionError, match="VERIFY FAIL"):
        verify_workbook(bad, c["rr"].inputs, c["rr"].policy)


def test_verify_catches_a_fragment_forged_out_of_a_complete_set():
    # review C1, forgery T1: one unit shipped OUT of a Complete full set —
    # the one law that never bends — once passed the verifier clean
    from restitch.core.verify import verify_workbook
    c = _bundle("tailoring")
    inputs, pol = c["rr"].inputs, c["rr"].policy
    line_keys = {(m["dst"], m["style"]) for m in c["R"]["moves"]}
    k, h = next((k, h) for k, h in inputs.holdings.items()
                if all(h.sizes.get(p, 0) > 0 for p in pol.pivotals)
                and k not in line_keys)
    dst = next(m["dst"] for m in c["R"]["moves"] if m["dst"] != k[0])
    piv = next(p for p in pol.pivotals if h.sizes.get(p, 0) > 0)
    bad = _forged(c, lambda ws, ix, hr, dr: _append_heal(
        ws, ix, dr, src=k[0], dst=dst, style=k[1], size=piv))
    with pytest.raises(AssertionError, match="v_full_set_freeze"):
        verify_workbook(bad, inputs, pol)


def test_verify_catches_a_ghost_size_shipped_from_zero():
    # review C1, forgery T2: a door ships a size it holds zero of
    from restitch.core.verify import verify_workbook
    c = _bundle("tailoring")
    inputs, pol = c["rr"].inputs, c["rr"].policy
    k, h, missing = next(
        (k, h, p) for k, h in inputs.holdings.items()
        for p in pol.pivotals if h.sizes.get(p, 0) <= 0)
    dst = next(m["dst"] for m in c["R"]["moves"] if m["dst"] != k[0])
    bad = _forged(c, lambda ws, ix, hr, dr: _append_heal(
        ws, ix, dr, src=k[0], dst=dst, style=k[1], size=missing))
    with pytest.raises(AssertionError, match="v_overdraw_size"):
        verify_workbook(bad, inputs, pol)


def test_verify_refuses_a_missing_required_header():
    from restitch.core.verify import verify_workbook
    c = _bundle("tailoring")
    bad = _forged(c, lambda ws, ix, hr, dr:
                  ws.cell(hr, ix["Qty"], "Quantity"))
    with pytest.raises(AssertionError, match="required headers missing"):
        verify_workbook(bad, c["rr"].inputs, c["rr"].policy)


# ── footwear-UK: the generalization gate ────────────────────────────────
def test_footwear_uk_runs_the_whole_pipeline():
    from restitch.core.verify import verify_workbook
    c = _bundle("footwear-uk")
    R = c["R"]
    assert sorted(R["categories"]) == ["Derby", "Loafer"]
    sizes_seen = {z for h in R["holdings"].values() for z in h.sizes}
    assert "6.5" in sizes_seen and "7" in sizes_seen, \
        "uk_shoe normalizer must fold raw floats (6.5 / 7.0) to canonical strings"
    assert not any(z.endswith(".0") for z in sizes_seen)
    assert len(R["moves"]) > 0
    vb = verify_workbook(c["out"], c["rr"].inputs, c["rr"].policy)
    assert vb.ok


def test_footwear_extremes_are_the_named_shoe_sizes():
    c = _bundle("footwear-uk")
    ext = [m for m in c["R"]["moves"] if m["program"] == "EXT"]
    assert all(m["size"] in ("6", "10") for m in ext)


if __name__ == "__main__":
    import sys
    mod = sys.modules["__main__"]
    for name in sorted(n for n in dir(mod) if n.startswith("test_")):
        getattr(mod, name)()
        print(f"{name}: OK")
